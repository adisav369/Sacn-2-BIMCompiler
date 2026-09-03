# BIM OOTB — Frictionless BIM. Two DBs. One browser. Zero install.
# Copyright (c) 2025-2026 Redhuan D. Oon <red1org@gmail.com>
# SPDX-License-Identifier: MIT
"""
nD BIM Engine — Template-Driven 4D/5D/6D/7D/8D Analytics
---------------------------------------------------------
Abstract engine: reads JSON templates, queries extracted DB, produces tables + Excel.
Zero hardcoded IFC classes. Users edit templates, engine applies formulas.

Usage:
    python nD_engine.py <database_path> [--dims 4D,5D,6D] [--templates /path/to/templates]
    python nD_engine.py <database_path> --all
    python nD_engine.py <database_path> --dims 5D --start-date 2025-06-01
    python nD_engine.py <database_path> --all --township
"""

import json
import logging
import sqlite3
import sys
import os
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# ---------------------------------------------------------------------------
# Logging — debug log to file, summary to console
# ---------------------------------------------------------------------------
LOG_DIR = Path(__file__).parent
LOG_FILE = LOG_DIR / 'nD_engine_log.txt'

logger = logging.getLogger('nD_engine')
logger.setLevel(logging.DEBUG)

# File handler — full debug
fh = logging.FileHandler(LOG_FILE, mode='w', encoding='utf-8')
fh.setLevel(logging.DEBUG)
fh.setFormatter(logging.Formatter('%(asctime)s [%(levelname)-5s] %(message)s', datefmt='%H:%M:%S'))
logger.addHandler(fh)

# Console handler — INFO and above
ch = logging.StreamHandler(sys.stdout)
ch.setLevel(logging.INFO)
ch.setFormatter(logging.Formatter('[%(levelname)-5s] %(message)s'))
logger.addHandler(ch)


# ---------------------------------------------------------------------------
# Template Loader
# ---------------------------------------------------------------------------
class TemplateLoader:
    """Loads and validates JSON templates from the templates directory."""

    def __init__(self, templates_dir: str = None):
        if templates_dir:
            self.templates_dir = Path(templates_dir)
        else:
            self.templates_dir = Path(__file__).parent.parent / 'templates'

        if not self.templates_dir.exists():
            raise FileNotFoundError(f"Templates directory not found: {self.templates_dir}")

        self._cache: Dict[str, dict] = {}
        logger.info("§TEMPLATES dir=%s", self.templates_dir)

    def load(self, filename: str) -> dict:
        """Load a JSON template, with caching."""
        if filename in self._cache:
            return self._cache[filename]

        path = self.templates_dir / filename
        if not path.exists():
            raise FileNotFoundError(f"Template not found: {path}")

        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        self._cache[filename] = data
        logger.info("§TEMPLATE_LOADED file=%s keys=%s", filename, list(data.keys()))
        return data

    def load_master(self) -> dict:
        """Load nD_formulas.json (the master template)."""
        return self.load('nD_formulas.json')


# ---------------------------------------------------------------------------
# DB Helpers
# ---------------------------------------------------------------------------
def has_column(conn: sqlite3.Connection, table: str, column: str) -> bool:
    """Check if a column exists in a table."""
    cols = conn.execute(f"PRAGMA table_info({table})").fetchall()
    return any(c[1] == column for c in cols)


def building_expr(conn: sqlite3.Connection) -> str:
    """Return SQL expression for building column, with fallback for old DBs."""
    if has_column(conn, 'elements_meta', 'building'):
        return 'building'
    return "'' AS building"


def building_expr_alias(conn: sqlite3.Connection, alias: str = 'e') -> str:
    """Return SQL expression for building column with table alias."""
    if has_column(conn, 'elements_meta', 'building'):
        return f'{alias}.building'
    return f"'' AS building"


def table_exists(conn: sqlite3.Connection, table_name: str) -> bool:
    row = conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (table_name,)
    ).fetchone()
    return row is not None


def get_element_count(conn: sqlite3.Connection) -> int:
    if not table_exists(conn, 'elements_meta'):
        return 0
    return conn.execute("SELECT COUNT(*) FROM elements_meta").fetchone()[0]


def get_ifc_classes(conn: sqlite3.Connection) -> List[Tuple[str, int]]:
    """Return [(ifc_class, count)] sorted by count descending."""
    return conn.execute("""
        SELECT ifc_class, COUNT(*) n FROM elements_meta
        WHERE ifc_class IS NOT NULL
        GROUP BY ifc_class ORDER BY n DESC
    """).fetchall()


# ---------------------------------------------------------------------------
# Township — archetype extraction + building awareness
# ---------------------------------------------------------------------------
import re

_SUBURB_RE = re.compile(r'^S\d+_\d+_(.+)$')
_CBD_RE = re.compile(r'^T0_(.+)$')


def extract_archetype(building: str) -> str:
    """Extract archetype name from building identifier.

    S0_0_HospitalGarage → HospitalGarage  (suburb tile)
    T0_Hospital         → Hospital         (CBD)
    Hospital            → Hospital         (bare name, single-building DB)
    """
    m = _SUBURB_RE.match(building)
    if m:
        return m.group(1)
    m = _CBD_RE.match(building)
    if m:
        return m.group(1)
    return building


def get_archetypes(conn: sqlite3.Connection) -> List[Tuple[str, int, int]]:
    """Return [(archetype, instance_count, total_elements)] for township mode."""
    if not has_column(conn, 'elements_meta', 'building'):
        count = conn.execute("SELECT COUNT(*) FROM elements_meta").fetchone()[0]
        return [('Unknown', 1, count)]
    rows = conn.execute("""
        SELECT building, COUNT(*) FROM elements_meta
        GROUP BY building ORDER BY building
    """).fetchall()

    arch_map: Dict[str, dict] = {}
    for building, count in rows:
        arch = extract_archetype(building)
        if arch not in arch_map:
            arch_map[arch] = {'instances': set(), 'elements': 0}
        arch_map[arch]['instances'].add(building)
        arch_map[arch]['elements'] += count

    result = [(arch, len(v['instances']), v['elements'])
              for arch, v in arch_map.items()]
    result.sort(key=lambda x: -x[2])
    return result


def get_buildings(conn: sqlite3.Connection) -> List[Tuple[str, int]]:
    """Return [(building, element_count)] sorted by count descending."""
    if not has_column(conn, 'elements_meta', 'building'):
        count = conn.execute("SELECT COUNT(*) FROM elements_meta").fetchone()[0]
        return [('', count)]
    return conn.execute("""
        SELECT building, COUNT(*) n FROM elements_meta
        GROUP BY building ORDER BY n DESC
    """).fetchall()


def is_multi_building(conn: sqlite3.Connection) -> bool:
    """Check if DB has more than one distinct building."""
    if not has_column(conn, 'elements_meta', 'building'):
        return False
    row = conn.execute("SELECT COUNT(DISTINCT building) FROM elements_meta").fetchone()
    return row[0] > 1


def resolve_measurement(ifc_class: str, rules: dict) -> str:
    """Resolve measurement type for an IFC class from nD_formulas measurement_rules."""
    for mtype in ('linear', 'area', 'volume'):
        if ifc_class in rules.get(mtype, []):
            return mtype
    return rules.get('_default', 'count')


# ---------------------------------------------------------------------------
# Ensure prerequisite tables exist (auto-run if missing)
# ---------------------------------------------------------------------------
def ensure_tables(conn: sqlite3.Connection, db_path: str, required: List[str],
                  loader: 'TemplateLoader', start_date: str):
    """Auto-generate prerequisite tables if missing."""
    for table in required:
        if table_exists(conn, table):
            logger.debug("§TABLE_EXISTS table=%s", table)
            continue

        logger.info("§TABLE_MISSING table=%s — auto-generating", table)
        if table == 'simple_qto':
            run_5d_qto(db_path, loader)
            conn.execute("SELECT 1")  # refresh connection state
        elif table == 'construction_schedule':
            run_4d_schedule(db_path, loader, start_date)
            conn.execute("SELECT 1")
        else:
            logger.warning("§TABLE_CANNOT_AUTO_CREATE table=%s — skipping", table)


# ---------------------------------------------------------------------------
# 5D — QTO + Cost (template-driven)
# ---------------------------------------------------------------------------
def run_5d_qto(db_path: str, loader: TemplateLoader):
    """Extract quantities and apply costs from 5D_rates.json template."""
    logger.info("=" * 70)
    logger.info("§5D_START — QTO + Cost Extraction")
    logger.info("=" * 70)

    master = loader.load_master()
    rates = loader.load('5D_rates.json')
    meas_rules = master['measurement_rules']

    conn = sqlite3.connect(db_path)
    has_bld = has_column(conn, 'elements_meta', 'building')
    bld_col = 'building' if has_bld else "'' AS building"
    bld_ref = 'building' if has_bld else "''"
    bld_eref = 'e.building' if has_bld else "'' AS building"
    bld_group = ', building' if has_bld else ''
    bld_egroup = ', e.building' if has_bld else ''
    total_elements = get_element_count(conn)
    logger.info("§5D_DB elements=%d db=%s has_building=%s", total_elements, Path(db_path).name, has_bld)

    # Drop and recreate
    conn.execute("DROP TABLE IF EXISTS simple_qto")
    conn.execute("""
    CREATE TABLE simple_qto (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        building TEXT,
        discipline TEXT,
        ifc_class TEXT,
        storey TEXT,
        measurement_type TEXT,
        element_count INTEGER,
        total_quantity REAL,
        uom TEXT,
        avg_quantity REAL,
        unit_cost_rm REAL,
        total_cost_rm REAL
    )""")
    logger.debug("§5D_TABLE_CREATED simple_qto")

    # Build SQL per measurement type
    has_rtree = table_exists(conn, 'elements_rtree')
    formulas = master['measurement_rules']['_formulas']

    ifc_classes = get_ifc_classes(conn)
    logger.info("§5D_IFC_CLASSES total_distinct=%d", len(ifc_classes))

    # Group IFC classes by measurement type
    groups = {'linear': [], 'area': [], 'volume': [], 'count': []}
    for ifc_class, count in ifc_classes:
        mtype = resolve_measurement(ifc_class, meas_rules)
        groups[mtype].append(ifc_class)
        logger.debug("§5D_MEASURE ifc=%s type=%s count=%d", ifc_class, mtype, count)

    # Insert per group
    for mtype, classes in groups.items():
        if not classes:
            continue

        placeholders = ','.join('?' * len(classes))

        if mtype == 'count' or not has_rtree:
            sql = f"""
            INSERT INTO simple_qto (building, discipline, ifc_class, storey, measurement_type,
                                    element_count, total_quantity, uom, avg_quantity)
            SELECT {bld_ref}, discipline, ifc_class, storey, 'COUNT',
                   COUNT(*), COUNT(*), 'EA', 1.0
            FROM elements_meta
            WHERE ifc_class IN ({placeholders})
            GROUP BY discipline, ifc_class, storey{bld_group}
            """
            uom = 'EA'
        elif mtype == 'linear':
            sql = f"""
            INSERT INTO simple_qto (building, discipline, ifc_class, storey, measurement_type,
                                    element_count, total_quantity, uom, avg_quantity)
            SELECT {bld_eref}, e.discipline, e.ifc_class, e.storey, 'LINEAR',
                   COUNT(*),
                   ROUND(SUM(r.maxZ - r.minZ), 2),
                   'M',
                   ROUND(AVG(r.maxZ - r.minZ), 2)
            FROM elements_meta e JOIN elements_rtree r ON e.id = r.id
            WHERE e.ifc_class IN ({placeholders})
            GROUP BY e.discipline, e.ifc_class, e.storey{bld_egroup}
            """
            uom = 'M'
        elif mtype == 'area':
            sql = f"""
            INSERT INTO simple_qto (building, discipline, ifc_class, storey, measurement_type,
                                    element_count, total_quantity, uom, avg_quantity)
            SELECT {bld_eref}, e.discipline, e.ifc_class, e.storey, 'AREA',
                   COUNT(*),
                   ROUND(SUM((r.maxX - r.minX) * (r.maxY - r.minY)), 2),
                   'M2',
                   ROUND(AVG((r.maxX - r.minX) * (r.maxY - r.minY)), 2)
            FROM elements_meta e JOIN elements_rtree r ON e.id = r.id
            WHERE e.ifc_class IN ({placeholders})
            GROUP BY e.discipline, e.ifc_class, e.storey{bld_egroup}
            """
            uom = 'M2'
        elif mtype == 'volume':
            sql = f"""
            INSERT INTO simple_qto (building, discipline, ifc_class, storey, measurement_type,
                                    element_count, total_quantity, uom, avg_quantity)
            SELECT {bld_eref}, e.discipline, e.ifc_class, e.storey, 'VOLUME',
                   COUNT(*),
                   ROUND(SUM((r.maxX-r.minX)*(r.maxY-r.minY)*(r.maxZ-r.minZ)), 2),
                   'M3',
                   ROUND(AVG((r.maxX-r.minX)*(r.maxY-r.minY)*(r.maxZ-r.minZ)), 2)
            FROM elements_meta e JOIN elements_rtree r ON e.id = r.id
            WHERE e.ifc_class IN ({placeholders})
            GROUP BY e.discipline, e.ifc_class, e.storey{bld_egroup}
            """
            uom = 'M3'

        conn.execute(sql, classes)
        logger.info("§5D_INSERT mtype=%s classes=%d uom=%s", mtype, len(classes), uom)

    conn.commit()

    # Apply costs from template
    material_rates = rates['material_rates']
    default_rate = material_rates.get('_default', {})
    costed = 0
    uncosted_classes = set()

    for row in conn.execute("SELECT DISTINCT ifc_class FROM simple_qto").fetchall():
        ifc_class = row[0]
        rate_entry = material_rates.get(ifc_class, default_rate)
        unit_rate = rate_entry.get('rate', 0)

        if unit_rate > 0:
            conn.execute("""
                UPDATE simple_qto SET unit_cost_rm = ?, total_cost_rm = ROUND(total_quantity * ?, 2)
                WHERE ifc_class = ?
            """, (unit_rate, unit_rate, ifc_class))
            costed += 1
            logger.debug("§5D_RATE ifc=%s rate=%.2f src=%s",
                         ifc_class, unit_rate,
                         'template' if ifc_class in material_rates else '_default')
        else:
            uncosted_classes.add(ifc_class)

    conn.commit()

    # Summary
    grand = conn.execute("SELECT SUM(total_cost_rm) FROM simple_qto WHERE total_cost_rm IS NOT NULL").fetchone()[0] or 0
    rows = conn.execute("SELECT COUNT(*) FROM simple_qto").fetchone()[0]
    conn.close()

    logger.info("§5D_DONE rows=%d costed=%d uncosted=%d grand_total=RM %.2f",
                rows, costed, len(uncosted_classes), grand)
    if uncosted_classes:
        logger.warning("§5D_UNCOSTED classes=%s", sorted(uncosted_classes))

    return rows


# ---------------------------------------------------------------------------
# 4D — Schedule (template-driven)
# ---------------------------------------------------------------------------
def run_4d_schedule(db_path: str, loader: TemplateLoader, start_date_str: str = '2025-01-06'):
    """Generate construction schedule from 4D_phases.json template."""
    logger.info("=" * 70)
    logger.info("§4D_START — Construction Schedule Generation")
    logger.info("=" * 70)

    phases_tmpl = loader.load('4D_phases.json')
    calendar = phases_tmpl['calendar']
    rules = phases_tmpl['ifc_class_rules']
    default_rule = rules.get('_default', {
        'phase': 'Unknown', 'sequence': 99, 'predecessors': [], 'resource': 'LABORER', 'productivity': 10.0
    })

    rates_tmpl = loader.load('5D_rates.json')
    labor_rates = rates_tmpl.get('labor_rates', {})
    equip_alloc = rates_tmpl.get('equipment_allocation', {})

    conn = sqlite3.connect(db_path)
    has_bld = has_column(conn, 'elements_meta', 'building')
    bld_ref = 'building' if has_bld else "'' AS building"
    bld_group = ', building' if has_bld else ''
    total_elements = get_element_count(conn)
    logger.info("§4D_DB elements=%d db=%s has_building=%s", total_elements, Path(db_path).name, has_bld)

    # Ensure schedule table
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS construction_schedule (
        task_id INTEGER PRIMARY KEY AUTOINCREMENT,
        wbs_code TEXT NOT NULL,
        task_name TEXT NOT NULL,
        building TEXT,
        element_guid TEXT,
        ifc_class TEXT,
        discipline TEXT,
        storey TEXT,
        phase TEXT,
        sequence INTEGER,
        quantity REAL,
        uom TEXT,
        productivity_rate REAL,
        duration_days REAL,
        start_date TEXT,
        finish_date TEXT,
        predecessors TEXT,
        labor_resource TEXT,
        labor_crew_size INTEGER,
        equipment_resource TEXT,
        status TEXT DEFAULT 'Not Started',
        percent_complete REAL DEFAULT 0.0,
        early_start TEXT, early_finish TEXT,
        late_start TEXT, late_finish TEXT,
        total_float_days REAL,
        is_critical BOOLEAN DEFAULT 0
    );
    CREATE INDEX IF NOT EXISTS idx_schedule_phase ON construction_schedule(phase);
    CREATE INDEX IF NOT EXISTS idx_schedule_discipline ON construction_schedule(discipline);
    CREATE INDEX IF NOT EXISTS idx_schedule_storey ON construction_schedule(storey);
    """)

    conn.execute("DELETE FROM construction_schedule")

    start_date = datetime.fromisoformat(start_date_str)
    working_days = set(calendar.get('working_days', [0, 1, 2, 3, 4, 5]))

    # Get activities grouped by (building, storey, discipline, ifc_class)
    activities = conn.execute(f"""
        SELECT {bld_ref}, COALESCE(storey, 'Unknown') as storey, discipline, ifc_class, COUNT(*) as cnt
        FROM elements_meta WHERE ifc_class IS NOT NULL
        GROUP BY storey, discipline, ifc_class{bld_group} ORDER BY storey, ifc_class
    """).fetchall()

    logger.info("§4D_ACTIVITIES total=%d", len(activities))

    tasks_created = 0
    phase_counts = {}
    unknown_count = 0

    for building, storey, discipline, ifc_class, element_count in activities:
        rule = rules.get(ifc_class, default_rule)
        phase = rule['phase']
        sequence = rule['sequence']
        resource = rule.get('resource', 'LABORER')
        productivity = rule.get('productivity', 10.0)

        if ifc_class not in rules:
            unknown_count += 1
            logger.debug("§4D_DEFAULT ifc=%s -> phase=%s (using _default)", ifc_class, phase)

        quantity = element_count
        duration_days = max(1.0, quantity / productivity)

        # Crew size from labor rates
        crew_size = labor_rates.get(resource, {}).get('crew_size', 2)

        # Equipment
        equipment = None
        if ifc_class in equip_alloc:
            equipment = equip_alloc[ifc_class].get('equipment')

        # Date calculation
        task_start = start_date + timedelta(days=sequence * 5)
        current = task_start
        remaining = duration_days
        while remaining > 0:
            if current.weekday() in working_days:
                remaining -= 1
            current += timedelta(days=1)
        task_finish = current

        bld_label = extract_archetype(building) if building else ''
        task_name = f"{bld_label} / {storey} - Install {ifc_class.replace('Ifc', '')} ({discipline})"
        wbs_code = f"1.{sequence}.{tasks_created}"

        conn.execute("""
            INSERT INTO construction_schedule (
                wbs_code, task_name, building, ifc_class, discipline, storey,
                phase, sequence, quantity, uom, productivity_rate, duration_days,
                start_date, finish_date, predecessors,
                labor_resource, labor_crew_size, equipment_resource,
                status, percent_complete
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (wbs_code, task_name, building, ifc_class, discipline, storey,
              phase, sequence, quantity, 'EA', productivity, duration_days,
              task_start.isoformat(), task_finish.isoformat(), '[]',
              resource, crew_size, equipment, 'Not Started', 0.0))

        tasks_created += 1
        phase_counts[phase] = phase_counts.get(phase, 0) + 1

        logger.debug("§4D_TASK [%s] %s qty=%d dur=%.1f days=%s->%s",
                      wbs_code, task_name, element_count, duration_days,
                      task_start.date(), task_finish.date())

    conn.commit()
    conn.close()

    logger.info("§4D_DONE tasks=%d unknown_fallback=%d", tasks_created, unknown_count)
    for phase, count in sorted(phase_counts.items(), key=lambda x: x[1], reverse=True):
        logger.info("§4D_PHASE %-20s tasks=%d", phase, count)

    return tasks_created


# ---------------------------------------------------------------------------
# 6D — Carbon Audit (template-driven)
# ---------------------------------------------------------------------------
def run_6d_carbon(db_path: str, loader: TemplateLoader):
    """Generate carbon audit from 6D_carbon.json template."""
    logger.info("=" * 70)
    logger.info("§6D_START — Sustainability / Carbon Audit")
    logger.info("=" * 70)

    carbon_tmpl = loader.load('6D_carbon.json')
    master = loader.load_master()
    factors = carbon_tmpl['carbon_factors']
    default_factor = factors.get('_default', {'factor': 50.0, 'per': 'EA'})
    meas_rules = master['measurement_rules']

    conn = sqlite3.connect(db_path)
    has_bld = has_column(conn, 'elements_meta', 'building')
    bld_ref = 'building' if has_bld else "'' AS building"
    bld_group = ', building' if has_bld else ''
    has_rtree = table_exists(conn, 'elements_rtree')

    conn.execute("DROP TABLE IF EXISTS carbon_audit")
    conn.execute("""
    CREATE TABLE carbon_audit (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        building TEXT,
        discipline TEXT,
        ifc_class TEXT,
        storey TEXT,
        element_count INTEGER,
        total_quantity REAL,
        uom TEXT,
        carbon_factor REAL,
        embodied_carbon_kg REAL,
        source_note TEXT
    )""")
    logger.debug("§6D_TABLE_CREATED carbon_audit")

    # If simple_qto exists, use it directly; otherwise compute from elements_meta
    if table_exists(conn, 'simple_qto'):
        logger.info("§6D_SOURCE using simple_qto table")
        rows = conn.execute("""
            SELECT building, discipline, ifc_class, storey, element_count, total_quantity, uom
            FROM simple_qto
        """).fetchall()
    else:
        logger.info("§6D_SOURCE using elements_meta (no simple_qto)")
        rows = conn.execute(f"""
            SELECT {bld_ref}, discipline, ifc_class, COALESCE(storey,'Unknown'), COUNT(*), COUNT(*), 'EA'
            FROM elements_meta WHERE ifc_class IS NOT NULL
            GROUP BY discipline, ifc_class, storey{bld_group}
        """).fetchall()

    total_carbon = 0.0
    for building, discipline, ifc_class, storey, elem_count, quantity, uom in rows:
        factor_entry = factors.get(ifc_class, default_factor)
        factor = factor_entry['factor']
        note = factor_entry.get('note', '')

        embodied = quantity * factor
        total_carbon += embodied

        conn.execute("""
            INSERT INTO carbon_audit (building, discipline, ifc_class, storey, element_count,
                                      total_quantity, uom, carbon_factor, embodied_carbon_kg, source_note)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (building, discipline, ifc_class, storey, elem_count, quantity, uom,
              factor, round(embodied, 2), note))

        logger.debug("§6D_ROW ifc=%s qty=%.1f factor=%.1f carbon=%.1f kg",
                      ifc_class, quantity, factor, embodied)

    conn.commit()
    row_count = conn.execute("SELECT COUNT(*) FROM carbon_audit").fetchone()[0]
    conn.close()

    logger.info("§6D_DONE rows=%d total_embodied_carbon=%.0f kgCO2e (%.1f tCO2e)",
                row_count, total_carbon, total_carbon / 1000)
    return row_count


# ---------------------------------------------------------------------------
# 7D — Asset Register / Lifecycle (template-driven)
# ---------------------------------------------------------------------------
def run_7d_lifecycle(db_path: str, loader: TemplateLoader):
    """Generate asset register from 7D_lifecycle.json template."""
    logger.info("=" * 70)
    logger.info("§7D_START — Facility Management / Asset Register")
    logger.info("=" * 70)

    lifecycle_tmpl = loader.load('7D_lifecycle.json')
    rules = lifecycle_tmpl['lifespan_rules']
    default_rule = rules.get('_default', {
        'lifespan_years': 30, 'warranty_months': 12, 'service_interval_months': 6
    })

    conn = sqlite3.connect(db_path)
    has_bld = has_column(conn, 'elements_meta', 'building')

    conn.execute("DROP TABLE IF EXISTS asset_register")
    conn.execute("""
    CREATE TABLE asset_register (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        building TEXT,
        guid TEXT,
        ifc_class TEXT,
        element_name TEXT,
        discipline TEXT,
        storey TEXT,
        material_name TEXT,
        asset_tag TEXT,
        lifespan_years INTEGER,
        warranty_months INTEGER,
        service_interval_months INTEGER,
        replacement_year INTEGER,
        lifecycle_note TEXT
    )""")
    logger.debug("§7D_TABLE_CREATED asset_register")

    # Per-element register (not grouped)
    bld_sel = 'building,' if has_bld else "'' AS building,"
    bld_order = 'building,' if has_bld else ''
    elements = conn.execute(f"""
        SELECT {bld_sel} guid, ifc_class, element_name, discipline, storey, material_name
        FROM elements_meta WHERE ifc_class IS NOT NULL
        ORDER BY {bld_order} discipline, storey, ifc_class
    """).fetchall()

    logger.info("§7D_ELEMENTS total=%d", len(elements))

    seq = {}  # per (building, discipline, storey) sequence counter for asset tags
    batch = []

    for building, guid, ifc_class, elem_name, discipline, storey, material in elements:
        rule = rules.get(ifc_class, default_rule)
        lifespan = rule['lifespan_years']
        warranty = rule['warranty_months']
        service = rule['service_interval_months']
        note = rule.get('note', '')
        replacement_yr = int(lifespan * 0.8)

        # Asset tag
        bld = extract_archetype(building) if building else 'UNK'
        key = (bld, discipline or 'UNK', storey or 'UNK')
        seq[key] = seq.get(key, 0) + 1
        asset_tag = f"{key[0]}_{key[1]}_{key[2]}_{seq[key]:04d}"

        batch.append((building, guid, ifc_class, elem_name, discipline, storey, material,
                       asset_tag, lifespan, warranty, service, replacement_yr, note))

    conn.executemany("""
        INSERT INTO asset_register (building, guid, ifc_class, element_name, discipline, storey,
                                    material_name, asset_tag, lifespan_years, warranty_months,
                                    service_interval_months, replacement_year, lifecycle_note)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, batch)

    conn.commit()
    row_count = conn.execute("SELECT COUNT(*) FROM asset_register").fetchone()[0]
    conn.close()

    logger.info("§7D_DONE assets=%d", row_count)
    return row_count


# ---------------------------------------------------------------------------
# 8D — Safety / Hazard Register (template-driven)
# ---------------------------------------------------------------------------
def run_8d_safety(db_path: str, loader: TemplateLoader):
    """Generate hazard register from 8D_safety.json template."""
    logger.info("=" * 70)
    logger.info("§8D_START — Safety / Hazard Register")
    logger.info("=" * 70)

    safety_tmpl = loader.load('8D_safety.json')
    phase_hazards = safety_tmpl['phase_hazards']
    ifc_hazards = safety_tmpl['ifc_class_hazards']
    risk_matrix = safety_tmpl['risk_matrix']
    default_hazard = ifc_hazards.get('_default', {
        'hazard_class': 'GENERAL', 'additional_ppe': [], 'permit_required': False
    })

    conn = sqlite3.connect(db_path)
    has_bld = has_column(conn, 'elements_meta', 'building')
    bld_ref = 'building' if has_bld else "'' AS building"
    bld_group = ', building' if has_bld else ''

    conn.execute("DROP TABLE IF EXISTS hazard_register")
    conn.execute("""
    CREATE TABLE hazard_register (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        building TEXT,
        ifc_class TEXT,
        discipline TEXT,
        storey TEXT,
        element_count INTEGER,
        phase TEXT,
        hazard_class TEXT,
        risk_level TEXT,
        hazards TEXT,
        ppe_required TEXT,
        additional_ppe TEXT,
        permit_required INTEGER
    )""")
    logger.debug("§8D_TABLE_CREATED hazard_register")

    # Need phase info — get from construction_schedule if available, else from 4D template
    phases_tmpl = loader.load('4D_phases.json')
    phase_rules = phases_tmpl['ifc_class_rules']
    default_phase_rule = phase_rules.get('_default', {'phase': 'Architecture'})

    activities = conn.execute(f"""
        SELECT {bld_ref}, ifc_class, discipline, COALESCE(storey,'Unknown'), COUNT(*)
        FROM elements_meta WHERE ifc_class IS NOT NULL
        GROUP BY ifc_class, discipline, storey{bld_group}
    """).fetchall()

    logger.info("§8D_ACTIVITIES total=%d", len(activities))

    for building, ifc_class, discipline, storey, count in activities:
        # Phase
        phase_rule = phase_rules.get(ifc_class, default_phase_rule)
        phase = phase_rule['phase']

        # Hazard from ifc_class
        hz = ifc_hazards.get(ifc_class, default_hazard)
        hazard_class = hz['hazard_class']
        additional_ppe = hz.get('additional_ppe', [])
        permit = hz.get('permit_required', False)

        # Phase-level hazards
        ph = phase_hazards.get(phase, {'risk_level': 'LOW', 'hazards': [], 'ppe': []})
        risk_level = ph['risk_level']
        hazards = ph.get('hazards', [])
        base_ppe = ph.get('ppe', [])

        # Override risk from risk matrix if class-specific
        rm = risk_matrix.get(hazard_class, {})
        if rm:
            risk_level = rm.get('risk', risk_level)

        # Merge PPE
        all_ppe = list(dict.fromkeys(base_ppe + additional_ppe))  # deduplicate, preserve order

        conn.execute("""
            INSERT INTO hazard_register (building, ifc_class, discipline, storey, element_count,
                                         phase, hazard_class, risk_level,
                                         hazards, ppe_required, additional_ppe, permit_required)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (building, ifc_class, discipline, storey, count,
              phase, hazard_class, risk_level,
              json.dumps(hazards), json.dumps(all_ppe), json.dumps(additional_ppe),
              1 if permit else 0))

        logger.debug("§8D_ROW ifc=%s phase=%s hazard=%s risk=%s permit=%s",
                      ifc_class, phase, hazard_class, risk_level, permit)

    conn.commit()
    row_count = conn.execute("SELECT COUNT(*) FROM hazard_register").fetchone()[0]
    conn.close()

    logger.info("§8D_DONE rows=%d", row_count)
    return row_count


# ---------------------------------------------------------------------------
# Orchestrator — run selected dimensions
# ---------------------------------------------------------------------------
DIMENSION_RUNNERS = {
    '4D': lambda db, loader, sd: run_4d_schedule(db, loader, sd),
    '5D': lambda db, loader, sd: run_5d_qto(db, loader),
    '6D': lambda db, loader, sd: run_6d_carbon(db, loader),
    '7D': lambda db, loader, sd: run_7d_lifecycle(db, loader),
    '8D': lambda db, loader, sd: run_8d_safety(db, loader),
}


def run_engine(db_path: str, dims: List[str] = None, templates_dir: str = None,
               start_date: str = '2025-01-06', township: bool = False):
    """
    Main entry point. Runs selected dimensions against an extracted DB.

    Args:
        db_path: Path to _extracted.db
        dims: List of dimensions to run, e.g. ['4D','5D','6D']. None = all enabled.
        templates_dir: Override templates directory
        start_date: Project start date for 4D scheduling
        township: If True, log archetype summary and dedup suburb tiles
    """
    logger.info("=" * 70)
    logger.info("§ENGINE_START nD BIM Engine v1.1 (township=%s)", township)
    logger.info("§ENGINE_DB %s", db_path)
    logger.info("§ENGINE_DATE %s", datetime.now().isoformat())
    logger.info("=" * 70)

    if not Path(db_path).exists():
        logger.error("§ENGINE_ERROR Database not found: %s", db_path)
        return

    loader = TemplateLoader(templates_dir)
    master = loader.load_master()

    # Resolve which dimensions to run
    all_dims = master['dimensions']
    if dims is None:
        dims = [k for k, v in all_dims.items() if v.get('enabled', True)]

    logger.info("§ENGINE_DIMS running=%s", dims)

    # Check DB has elements_meta
    conn = sqlite3.connect(db_path)
    elem_count = get_element_count(conn)
    if elem_count == 0:
        logger.error("§ENGINE_ERROR elements_meta is empty or missing")
        conn.close()
        return

    ifc_classes = get_ifc_classes(conn)
    logger.info("§ENGINE_ELEMENTS total=%d distinct_ifc_classes=%d", elem_count, len(ifc_classes))
    for cls, cnt in ifc_classes[:10]:
        logger.debug("§ENGINE_CLASS %-30s %d", cls, cnt)
    if len(ifc_classes) > 10:
        logger.debug("§ENGINE_CLASS ... and %d more classes", len(ifc_classes) - 10)

    # Township auto-detect: >1 building = township mode (override with --township)
    multi = is_multi_building(conn)
    if multi and not township:
        township = True
        logger.info("§TOWNSHIP_AUTO multi-building DB detected — enabling township mode")

    if township:
        archetypes = get_archetypes(conn)
        buildings = get_buildings(conn)
        logger.info("§TOWNSHIP buildings=%d archetypes=%d", len(buildings), len(archetypes))
        for arch, inst, elems in archetypes:
            logger.info("§TOWNSHIP_ARCH %-30s instances=%d elements=%d", arch, inst, elems)

    conn.close()

    # Run dimensions in dependency order
    # 5D must run before 6D (needs simple_qto). 4D must run before 8D (needs schedule).
    order = ['5D', '4D', '6D', '7D', '8D']
    sorted_dims = [d for d in order if d in dims]

    results = {}
    for dim in sorted_dims:
        dim_info = all_dims.get(dim, {})
        logger.info("")
        logger.info(">>> Running %s — %s", dim, dim_info.get('name', dim))

        # Ensure prerequisites
        conn = sqlite3.connect(db_path)
        required = dim_info.get('requires_tables', [])
        ensure_tables(conn, db_path, required, loader, start_date)
        conn.close()

        # Run
        runner = DIMENSION_RUNNERS.get(dim)
        if runner:
            try:
                result = runner(db_path, loader, start_date)
                results[dim] = result
                logger.info("§ENGINE_RESULT %s = %s", dim, result)
            except Exception as e:
                logger.error("§ENGINE_ERROR %s failed: %s", dim, e, exc_info=True)
                results[dim] = f"ERROR: {e}"
        else:
            logger.warning("§ENGINE_SKIP %s — no runner implemented", dim)

    # Final summary
    logger.info("")
    logger.info("=" * 70)
    logger.info("§ENGINE_DONE — Summary")
    logger.info("=" * 70)
    for dim, result in results.items():
        logger.info("  %s: %s", dim, result)

    # Township Excel export
    if township and ('5D' in sorted_dims or '4D' in sorted_dims):
        try:
            excel_path = export_township_excel(db_path)
            results['EXCEL'] = excel_path
            logger.info("§ENGINE_EXCEL %s", excel_path)
        except Exception as e:
            logger.error("§ENGINE_EXCEL_ERROR %s", e, exc_info=True)

    logger.info("§ENGINE_LOG %s", LOG_FILE)

    return results


# ---------------------------------------------------------------------------
# Township Excel Export — per-archetype sheets + executive summary
# ---------------------------------------------------------------------------
def export_township_excel(db_path: str) -> str:
    """Generate township Excel workbook with per-archetype BOQ + schedule sheets.

    Reads building-aware simple_qto and construction_schedule tables.
    Deduplicates suburb tiles: compute once per archetype, multiply by instance count.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("§EXCEL_ERROR openpyxl not installed — pip install openpyxl")
        raise

    conn = sqlite3.connect(db_path)
    archetypes = get_archetypes(conn)

    logger.info("§EXCEL_START township export — %d archetypes", len(archetypes))

    wb = Workbook()
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    currency_fmt = '#,##0.00'
    number_fmt = '#,##0'
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, color='FFFFFF', size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # --- Cover Sheet ---
    ws_cover = wb.active
    ws_cover.title = 'Cover Sheet'
    ws_cover['A1'] = 'TOWNSHIP nD REPORT'
    ws_cover['A1'].font = Font(bold=True, size=18)
    ws_cover['A3'] = 'Database'
    ws_cover['B3'] = Path(db_path).name
    ws_cover['A4'] = 'Generated'
    ws_cover['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M')
    ws_cover['A5'] = 'Total Elements'
    ws_cover['B5'] = get_element_count(conn)
    ws_cover['A6'] = 'Buildings'
    ws_cover['B6'] = len(get_buildings(conn))
    ws_cover['A7'] = 'Archetypes'
    ws_cover['B7'] = len(archetypes)

    row = 9
    ws_cover.cell(row=row, column=1, value='Archetype').font = header_font
    ws_cover.cell(row=row, column=2, value='Instances').font = header_font
    ws_cover.cell(row=row, column=3, value='Elements').font = header_font
    for arch, inst, elems in archetypes:
        row += 1
        ws_cover.cell(row=row, column=1, value=arch)
        ws_cover.cell(row=row, column=2, value=inst)
        ws_cover.cell(row=row, column=3, value=elems)

    ws_cover.column_dimensions['A'].width = 30
    ws_cover.column_dimensions['B'].width = 20
    ws_cover.column_dimensions['C'].width = 15

    # --- Executive Summary ---
    ws_exec = wb.create_sheet('Executive Summary')

    has_qto = table_exists(conn, 'simple_qto')
    if has_qto:
        # Per-archetype cost summary (first instance only, then × instance count)
        ws_exec['A1'] = 'TOWNSHIP COST SUMMARY'
        ws_exec['A1'].font = title_font

        def _write_header(ws, row, cols):
            for c, label in enumerate(cols, 1):
                cell = ws.cell(row=row, column=c, value=label)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.border = thin_border

        _write_header(ws_exec, 3, ['Archetype', 'Instances', 'Archetype Cost (RM)',
                                    'Total Cost (RM)', 'Elements'])
        r = 4
        grand_total = 0.0
        for arch, inst, elems in archetypes:
            # Get cost for ONE instance of this archetype (any single building)
            # Pick buildings matching this archetype
            buildings_for_arch = conn.execute("""
                SELECT DISTINCT building FROM elements_meta
            """).fetchall()
            # Filter to first building matching archetype
            sample_building = None
            for (b,) in buildings_for_arch:
                if extract_archetype(b) == arch:
                    sample_building = b
                    break

            if sample_building:
                arch_cost_row = conn.execute("""
                    SELECT COALESCE(SUM(total_cost_rm), 0)
                    FROM simple_qto WHERE building = ? AND total_cost_rm IS NOT NULL
                """, (sample_building,)).fetchone()
                arch_cost = arch_cost_row[0]
            else:
                arch_cost = 0.0

            total_cost = arch_cost * inst
            grand_total += total_cost

            ws_exec.cell(row=r, column=1, value=arch).border = thin_border
            ws_exec.cell(row=r, column=2, value=inst).border = thin_border
            c = ws_exec.cell(row=r, column=3, value=arch_cost)
            c.number_format = currency_fmt
            c.border = thin_border
            c = ws_exec.cell(row=r, column=4, value=total_cost)
            c.number_format = currency_fmt
            c.border = thin_border
            ws_exec.cell(row=r, column=5, value=elems).border = thin_border
            r += 1

        # Grand total row
        r += 1
        ws_exec.cell(row=r, column=1, value='GRAND TOTAL').font = header_font
        c = ws_exec.cell(row=r, column=4, value=grand_total)
        c.number_format = currency_fmt
        c.font = header_font

        for col in range(1, 6):
            ws_exec.column_dimensions[get_column_letter(col)].width = 22

        logger.info("§EXCEL_EXEC grand_total=RM %.2f archetypes=%d", grand_total, len(archetypes))

    # --- Per-archetype BOQ sheets ---
    if has_qto:
        for arch, inst, elems in archetypes:
            sheet_name = f'BOQ - {arch}'[:31]  # Excel 31-char limit
            ws = wb.create_sheet(sheet_name)

            ws['A1'] = f'{arch} — Bill of Quantities'
            ws['A1'].font = title_font
            ws['A2'] = f'Instances: {inst} | Elements per instance: {elems // inst}'

            # Get sample building for this archetype
            sample = None
            for (b,) in conn.execute("SELECT DISTINCT building FROM elements_meta").fetchall():
                if extract_archetype(b) == arch:
                    sample = b
                    break

            if not sample:
                continue

            _write_header(ws, 4, ['Discipline', 'IFC Class', 'Storey', 'Qty (1×)',
                                   'UOM', 'Unit Rate (RM)', 'Cost (1×)', f'Cost (×{inst})'])

            rows = conn.execute("""
                SELECT discipline, ifc_class, storey, element_count,
                       uom, COALESCE(unit_cost_rm, 0), COALESCE(total_cost_rm, 0)
                FROM simple_qto WHERE building = ?
                ORDER BY discipline, ifc_class, storey
            """, (sample,)).fetchall()

            r = 5
            for disc, ifc, storey, qty, uom, unit_rate, cost_1x in rows:
                ws.cell(row=r, column=1, value=disc).border = thin_border
                ws.cell(row=r, column=2, value=ifc).border = thin_border
                ws.cell(row=r, column=3, value=storey).border = thin_border
                ws.cell(row=r, column=4, value=qty).border = thin_border
                ws.cell(row=r, column=5, value=uom).border = thin_border
                c = ws.cell(row=r, column=6, value=unit_rate)
                c.number_format = currency_fmt
                c.border = thin_border
                c = ws.cell(row=r, column=7, value=cost_1x)
                c.number_format = currency_fmt
                c.border = thin_border
                c = ws.cell(row=r, column=8, value=cost_1x * inst)
                c.number_format = currency_fmt
                c.border = thin_border
                r += 1

            # Total row
            r += 1
            arch_total = sum(row[6] for row in rows)
            ws.cell(row=r, column=6, value='TOTAL').font = header_font
            c = ws.cell(row=r, column=7, value=arch_total)
            c.number_format = currency_fmt
            c.font = header_font
            c = ws.cell(row=r, column=8, value=arch_total * inst)
            c.number_format = currency_fmt
            c.font = header_font

            for col in range(1, 9):
                ws.cell(row=r, column=col).border = thin_border
                ws.column_dimensions[get_column_letter(col)].width = 18

            logger.info("§EXCEL_BOQ arch=%s rows=%d cost_1x=RM %.2f cost_total=RM %.2f",
                         arch, len(rows), arch_total, arch_total * inst)

    # --- Schedule summary sheet ---
    has_schedule = table_exists(conn, 'construction_schedule')
    if has_schedule:
        ws_sched = wb.create_sheet('Schedule - Township')

        ws_sched['A1'] = 'TOWNSHIP CONSTRUCTION SCHEDULE'
        ws_sched['A1'].font = title_font

        _write_header(ws_sched, 3, ['Archetype', 'Phase', 'Tasks', 'Total Duration (days)',
                                     'Earliest Start', 'Latest Finish'])

        # Aggregate by archetype + phase
        sched_rows = conn.execute("""
            SELECT building, phase, COUNT(*), SUM(duration_days),
                   MIN(start_date), MAX(finish_date)
            FROM construction_schedule
            GROUP BY building, phase
            ORDER BY building, MIN(sequence)
        """).fetchall()

        # Dedup to archetypes (use first instance per archetype)
        seen_archetypes = {}
        for building, phase, tasks, dur, start, finish in sched_rows:
            arch = extract_archetype(building)
            key = (arch, phase)
            if key not in seen_archetypes:
                seen_archetypes[key] = (arch, phase, tasks, dur, start, finish)

        r = 4
        for key in sorted(seen_archetypes.keys()):
            arch, phase, tasks, dur, start, finish = seen_archetypes[key]
            ws_sched.cell(row=r, column=1, value=arch).border = thin_border
            ws_sched.cell(row=r, column=2, value=phase).border = thin_border
            ws_sched.cell(row=r, column=3, value=tasks).border = thin_border
            c = ws_sched.cell(row=r, column=4, value=round(dur, 1))
            c.number_format = number_fmt
            c.border = thin_border
            ws_sched.cell(row=r, column=5, value=start).border = thin_border
            ws_sched.cell(row=r, column=6, value=finish).border = thin_border
            r += 1

        for col in range(1, 7):
            ws_sched.column_dimensions[get_column_letter(col)].width = 22

        logger.info("§EXCEL_SCHED rows=%d", len(seen_archetypes))

    conn.close()

    # Save
    output_dir = Path(db_path).parent.parent / 'boq_reports'
    output_dir.mkdir(parents=True, exist_ok=True)
    db_stem = Path(db_path).stem.replace('_extracted', '')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = output_dir / f'Township_{db_stem}_{timestamp}.xlsx'
    wb.save(str(output_path))

    logger.info("§EXCEL_DONE %s (%d sheets)", output_path, len(wb.sheetnames))
    return str(output_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main():
    import argparse

    parser = argparse.ArgumentParser(
        description='nD BIM Engine — Template-Driven 4D/5D/6D/7D/8D Analytics',
        epilog='Example: python nD_engine.py myproject_extracted.db --all'
    )
    parser.add_argument('database', help='Path to _extracted.db')
    parser.add_argument('--dims', type=str, default=None,
                        help='Comma-separated dimensions to run: 4D,5D,6D,7D,8D')
    parser.add_argument('--all', action='store_true', help='Run all enabled dimensions')
    parser.add_argument('--templates', type=str, default=None,
                        help='Path to templates directory (default: ../templates/)')
    parser.add_argument('--start-date', type=str, default='2025-01-06',
                        help='Project start date for 4D (default: 2025-01-06)')
    parser.add_argument('--township', action='store_true',
                        help='Township mode: per-building breakdown, archetype dedup')

    args = parser.parse_args()

    dims = None
    if args.dims:
        dims = [d.strip().upper() for d in args.dims.split(',')]
    elif args.all:
        dims = None  # run_engine defaults to all enabled

    run_engine(args.database, dims=dims, templates_dir=args.templates,
               start_date=args.start_date, township=args.township)


if __name__ == '__main__':
    main()
