"""
Excel Schedule Exporter
-----------------------
Convert construction schedule to Excel format for users without MS Project.

Creates a Gantt-style spreadsheet with:
- Task list with dates and durations
- Resource allocation
- Visual timeline (conditional formatting)
- Summary statistics
"""

import sqlite3
import os
from datetime import datetime, timedelta
from pathlib import Path


def export_schedule_to_excel(db_path: str, output_path: str = None, project_name: str = "Terminal 1 Construction") -> str:
    """
    Export construction_schedule table to Excel with Gantt visualization.

    Args:
        db_path: Path to federation database with construction_schedule table
        output_path: Output Excel file path (auto-generated if None)
        project_name: Project name for Excel file

    Returns:
        Path to generated Excel file
    """
    # Check openpyxl availability
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import PieChart, BarChart, LineChart, ScatterChart, Reference, Series
        from openpyxl.chart.marker import Marker
    except ImportError:
        raise ImportError("openpyxl required for Excel export. Install: pip install openpyxl")

    if not Path(db_path).exists():
        raise FileNotFoundError(f"Database not found: {db_path}")

    # Auto-generate output path if not provided
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"Terminal1_Schedule_{timestamp}.xlsx"

    # Connect to database
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Check if construction_schedule table exists
    cursor.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='construction_schedule'"
    )
    if not cursor.fetchone():
        raise ValueError("construction_schedule table not found. Run schedule generator first.")

    # Read all tasks
    cursor.execute("""
        SELECT
            task_id, wbs_code, task_name, phase, sequence,
            quantity, uom, productivity_rate, duration_days,
            start_date, finish_date,
            labor_resource, labor_crew_size, equipment_resource,
            status, percent_complete,
            discipline, storey
        FROM construction_schedule
        ORDER BY sequence, task_id
    """)

    tasks = cursor.fetchall()
    conn.close()

    if not tasks:
        raise ValueError("No tasks found in construction_schedule table.")

    print(f"\n{'='*80}")
    print(f"EXCEL SCHEDULE EXPORT")
    print(f"{'='*80}")
    print(f"Database: {Path(db_path).name}")
    print(f"Tasks: {len(tasks)}")
    print(f"Output: {output_path}")
    print(f"{'='*80}\n")

    # Create workbook
    wb = Workbook()

    # Sheet 1: Schedule
    ws_schedule = wb.active
    ws_schedule.title = "Construction Schedule"

    # Header row
    headers = [
        "WBS", "Task Name", "Phase", "Discipline", "Storey",
        "Quantity", "UOM", "Productivity", "Duration (days)",
        "Start Date", "Finish Date", "Labor Resource", "Crew Size",
        "Equipment", "Status", "% Complete"
    ]

    # Style header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws_schedule.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Write tasks
    for row_idx, task_row in enumerate(tasks, start=2):
        (task_id, wbs_code, task_name, phase, sequence,
         quantity, uom, productivity, duration_days,
         start_date, finish_date,
         labor_resource, labor_crew_size, equipment_resource,
         status, percent_complete,
         discipline, storey) = task_row

        row_data = [
            wbs_code,
            task_name,
            phase,
            discipline,
            storey,
            quantity,
            uom,
            productivity,
            duration_days,
            start_date,
            finish_date,
            labor_resource,
            labor_crew_size,
            equipment_resource,
            status,
            percent_complete
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws_schedule.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border

            # Format duration with 2 decimal points
            if col_idx == 9:  # Duration column
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal='right')
            # Align numbers right
            elif col_idx in [6, 8, 13, 16]:
                cell.alignment = Alignment(horizontal='right')

        print(f"  [{wbs_code}] {task_name} | {duration_days:.2f} days | {start_date} → {finish_date}")

    # Auto-size columns
    for col_idx in range(1, len(headers) + 1):
        ws_schedule.column_dimensions[get_column_letter(col_idx)].width = 15

    # Widen task name column
    ws_schedule.column_dimensions['B'].width = 40

    # Sheet 2: Summary with Charts
    ws_summary = wb.create_sheet("Project Summary")

    # Calculate project duration
    if tasks:
        start_date = datetime.fromisoformat(tasks[0][9])
        finish_date = datetime.fromisoformat(tasks[-1][10])
        total_duration = (finish_date - start_date).days + 1
    else:
        total_duration = 0

    # Project info header
    title_cell = ws_summary.cell(row=1, column=1)
    title_cell.value = project_name
    title_cell.font = Font(bold=True, size=14, color="366092")

    # Project statistics
    info_data = [
        ["Database", Path(db_path).name],
        ["Total Tasks", len(tasks)],
        ["Project Start", tasks[0][9] if tasks else "N/A"],
        ["Project Finish", tasks[-1][10] if tasks else "N/A"],
        ["Total Duration", f"{total_duration} days"],
        ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]

    for idx, (label, value) in enumerate(info_data, start=2):
        ws_summary.cell(row=idx, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=idx, column=2, value=value)

    # Count tasks by phase
    phase_counts = {}
    for task in tasks:
        phase = task[3]
        phase_counts[phase] = phase_counts.get(phase, 0) + 1

    # Phase summary section
    phase_start_row = len(info_data) + 4
    header = ws_summary.cell(row=phase_start_row, column=1)
    header.value = "Phase Summary"
    header.font = Font(bold=True, size=12)
    header.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header.font = Font(bold=True, color="FFFFFF")

    ws_summary.cell(row=phase_start_row, column=2, value="Task Count").font = Font(bold=True, color="FFFFFF")
    ws_summary.cell(row=phase_start_row, column=2).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    for idx, (phase, count) in enumerate(sorted(phase_counts.items()), start=1):
        ws_summary.cell(row=phase_start_row + idx, column=1, value=phase)
        ws_summary.cell(row=phase_start_row + idx, column=2, value=count)

    # Count tasks by discipline
    discipline_counts = {}
    for task in tasks:
        discipline = task[16]
        discipline_counts[discipline] = discipline_counts.get(discipline, 0) + 1

    # Discipline summary section
    disc_start_row = phase_start_row + len(phase_counts) + 3
    header = ws_summary.cell(row=disc_start_row, column=1)
    header.value = "Discipline Summary"
    header.font = Font(bold=True, size=12)
    header.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header.font = Font(bold=True, color="FFFFFF")

    ws_summary.cell(row=disc_start_row, column=2, value="Task Count").font = Font(bold=True, color="FFFFFF")
    ws_summary.cell(row=disc_start_row, column=2).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    for idx, (discipline, count) in enumerate(sorted(discipline_counts.items()), start=1):
        ws_summary.cell(row=disc_start_row + idx, column=1, value=discipline)
        ws_summary.cell(row=disc_start_row + idx, column=2, value=count)

    # Column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20

    # Add Pie Chart - Phase Distribution
    pie_phase = PieChart()
    pie_phase.title = "Task Distribution by Phase"
    pie_phase.style = 10
    pie_phase.height = 10
    pie_phase.width = 15

    # Data references
    labels = Reference(ws_summary, min_col=1, min_row=phase_start_row + 1, max_row=phase_start_row + len(phase_counts))
    data = Reference(ws_summary, min_col=2, min_row=phase_start_row, max_row=phase_start_row + len(phase_counts))
    pie_phase.add_data(data, titles_from_data=True)
    pie_phase.set_categories(labels)

    # Position chart
    ws_summary.add_chart(pie_phase, 'D2')

    # Add Bar Chart - Discipline Distribution
    bar_disc = BarChart()
    bar_disc.title = "Task Count by Discipline"
    bar_disc.style = 10
    bar_disc.height = 10
    bar_disc.width = 15
    bar_disc.y_axis.title = "Number of Tasks"
    bar_disc.x_axis.title = "Discipline"

    # Data references
    labels = Reference(ws_summary, min_col=1, min_row=disc_start_row + 1, max_row=disc_start_row + len(discipline_counts))
    data = Reference(ws_summary, min_col=2, min_row=disc_start_row, max_row=disc_start_row + len(discipline_counts))
    bar_disc.add_data(data, titles_from_data=True)
    bar_disc.set_categories(labels)

    # Position chart
    ws_summary.add_chart(bar_disc, 'D18')

    # ========================================================================
    # Sheet 3: BIM 5D Dashboard (Comprehensive Project Analytics)
    # ========================================================================
    ws_dashboard = wb.create_sheet("BIM 5D Dashboard")

    # Dashboard title
    title = ws_dashboard.cell(row=1, column=1)
    title.value = f"{project_name} - BIM 5D Analytics Dashboard"
    title.font = Font(bold=True, size=16, color="366092")
    ws_dashboard.merge_cells('A1:F1')
    title.alignment = Alignment(horizontal='center')

    # ========================================================================
    # Chart 1: Phase Duration Analysis (Horizontal Bar)
    # ========================================================================
    chart1_start_row = 3
    ws_dashboard.cell(row=chart1_start_row, column=1, value="Phase Duration Analysis").font = Font(bold=True, size=11)
    ws_dashboard.cell(row=chart1_start_row + 1, column=1, value="Phase").font = Font(bold=True)
    ws_dashboard.cell(row=chart1_start_row + 1, column=2, value="Total Days").font = Font(bold=True)

    # Calculate total duration per phase
    phase_durations = {}
    for task in tasks:
        phase = task[3]
        duration = task[8]
        phase_durations[phase] = phase_durations.get(phase, 0) + duration

    row = chart1_start_row + 2
    for phase, total_days in sorted(phase_durations.items(), key=lambda x: x[1], reverse=True):
        ws_dashboard.cell(row=row, column=1, value=phase)
        ws_dashboard.cell(row=row, column=2, value=round(total_days, 2))
        row += 1

    # Create horizontal bar chart
    chart_phase_dur = BarChart()
    chart_phase_dur.type = "bar"
    chart_phase_dur.title = "Phase Duration (Total Days)"
    chart_phase_dur.style = 11
    chart_phase_dur.height = 8
    chart_phase_dur.width = 14
    chart_phase_dur.y_axis.title = "Phase"
    chart_phase_dur.x_axis.title = "Days"

    data = Reference(ws_dashboard, min_col=2, min_row=chart1_start_row + 1, max_row=chart1_start_row + 1 + len(phase_durations))
    categories = Reference(ws_dashboard, min_col=1, min_row=chart1_start_row + 2, max_row=chart1_start_row + 1 + len(phase_durations))
    chart_phase_dur.add_data(data, titles_from_data=True)
    chart_phase_dur.set_categories(categories)
    ws_dashboard.add_chart(chart_phase_dur, 'D3')

    # ========================================================================
    # Chart 2: Resource Workload by Phase
    # ========================================================================
    chart2_start_row = chart1_start_row + len(phase_durations) + 4
    ws_dashboard.cell(row=chart2_start_row, column=1, value="Resource Workload Analysis").font = Font(bold=True, size=11)
    ws_dashboard.cell(row=chart2_start_row + 1, column=1, value="Phase").font = Font(bold=True)
    ws_dashboard.cell(row=chart2_start_row + 1, column=2, value="Total Man-Days").font = Font(bold=True)

    # Calculate labor man-days per phase (crew_size * duration)
    phase_mandays = {}
    for task in tasks:
        phase = task[3]
        duration = task[8]
        crew_size = task[12] if task[12] else 1
        mandays = duration * crew_size
        phase_mandays[phase] = phase_mandays.get(phase, 0) + mandays

    row = chart2_start_row + 2
    for phase, mandays in sorted(phase_mandays.items(), key=lambda x: x[1], reverse=True):
        ws_dashboard.cell(row=row, column=1, value=phase)
        ws_dashboard.cell(row=row, column=2, value=round(mandays, 1))
        row += 1

    # Create column chart
    chart_resource = BarChart()
    chart_resource.type = "col"
    chart_resource.title = "Resource Workload (Man-Days by Phase)"
    chart_resource.style = 12
    chart_resource.height = 8
    chart_resource.width = 14
    chart_resource.x_axis.title = "Phase"
    chart_resource.y_axis.title = "Man-Days"

    data = Reference(ws_dashboard, min_col=2, min_row=chart2_start_row + 1, max_row=chart2_start_row + 1 + len(phase_mandays))
    categories = Reference(ws_dashboard, min_col=1, min_row=chart2_start_row + 2, max_row=chart2_start_row + 1 + len(phase_mandays))
    chart_resource.add_data(data, titles_from_data=True)
    chart_resource.set_categories(categories)
    ws_dashboard.add_chart(chart_resource, 'D16')

    # ========================================================================
    # Chart 3: S-Curve Progress (Cumulative Task Completion)
    # ========================================================================
    chart3_start_row = chart2_start_row + len(phase_mandays) + 4
    ws_dashboard.cell(row=chart3_start_row, column=1, value="S-Curve Progress Analysis").font = Font(bold=True, size=11)
    ws_dashboard.cell(row=chart3_start_row + 1, column=1, value="Week").font = Font(bold=True)
    ws_dashboard.cell(row=chart3_start_row + 1, column=2, value="Cumulative %").font = Font(bold=True)

    # Calculate cumulative completion over time (weekly intervals)
    if tasks:
        project_start = datetime.fromisoformat(tasks[0][9])
        project_end = datetime.fromisoformat(tasks[-1][10])
        total_weeks = max(1, ((project_end - project_start).days // 7) + 1)

        # Create weekly buckets
        weeks = []
        cumulative_pct = []
        for week_num in range(0, total_weeks + 1):
            week_end_date = project_start + timedelta(weeks=week_num)

            # Count tasks completed by this date
            completed = sum(1 for t in tasks if datetime.fromisoformat(t[10]) <= week_end_date)
            pct = (completed / len(tasks)) * 100

            weeks.append(f"W{week_num}")
            cumulative_pct.append(round(pct, 1))

        # Write data (limit to 20 points for readability)
        sample_interval = max(1, len(weeks) // 20)
        row = chart3_start_row + 2
        for i in range(0, len(weeks), sample_interval):
            ws_dashboard.cell(row=row, column=1, value=weeks[i])
            ws_dashboard.cell(row=row, column=2, value=cumulative_pct[i])
            row += 1

        # Add final point
        if (len(weeks) - 1) % sample_interval != 0:
            ws_dashboard.cell(row=row, column=1, value=weeks[-1])
            ws_dashboard.cell(row=row, column=2, value=cumulative_pct[-1])
            row += 1

        # Create S-curve line chart
        chart_scurve = LineChart()
        chart_scurve.title = "Project S-Curve (Cumulative Task Completion)"
        chart_scurve.style = 13
        chart_scurve.height = 8
        chart_scurve.width = 14
        chart_scurve.x_axis.title = "Project Week"
        chart_scurve.y_axis.title = "% Complete"
        chart_scurve.y_axis.scaling.min = 0
        chart_scurve.y_axis.scaling.max = 100

        data_rows = row - chart3_start_row - 2
        data = Reference(ws_dashboard, min_col=2, min_row=chart3_start_row + 1, max_row=chart3_start_row + 1 + data_rows)
        categories = Reference(ws_dashboard, min_col=1, min_row=chart3_start_row + 2, max_row=chart3_start_row + 1 + data_rows)
        chart_scurve.add_data(data, titles_from_data=True)
        chart_scurve.set_categories(categories)
        ws_dashboard.add_chart(chart_scurve, 'D29')

    # ========================================================================
    # Chart 4: Milestone Timeline (Phase Start/Finish Dates)
    # ========================================================================
    chart4_start_row = chart3_start_row + 25
    ws_dashboard.cell(row=chart4_start_row, column=1, value="Milestone Timeline").font = Font(bold=True, size=11)
    ws_dashboard.cell(row=chart4_start_row + 1, column=1, value="Milestone").font = Font(bold=True)
    ws_dashboard.cell(row=chart4_start_row + 1, column=2, value="Day Offset").font = Font(bold=True)
    ws_dashboard.cell(row=chart4_start_row + 1, column=3, value="Date").font = Font(bold=True)

    # Extract phase start/end dates
    phase_dates = {}
    for task in tasks:
        phase = task[3]
        start = task[9]
        finish = task[10]

        if phase not in phase_dates:
            phase_dates[phase] = {"start": start, "finish": finish}
        else:
            if start < phase_dates[phase]["start"]:
                phase_dates[phase]["start"] = start
            if finish > phase_dates[phase]["finish"]:
                phase_dates[phase]["finish"] = finish

    # Create milestone list
    milestones = []
    milestones.append(("🏁 Project Start", tasks[0][9]))
    for phase in sorted(phase_dates.keys()):
        milestones.append((f"▶ {phase} Start", phase_dates[phase]["start"]))
        milestones.append((f"◀ {phase} End", phase_dates[phase]["finish"]))
    milestones.append(("🏁 Project Finish", tasks[-1][10]))

    # Remove duplicates
    milestones = list(dict.fromkeys(milestones))

    # Write milestones with day offset from project start
    row = chart4_start_row + 2
    for milestone_name, date_str in sorted(milestones, key=lambda x: x[1]):
        date_obj = datetime.fromisoformat(date_str)
        day_offset = (date_obj - datetime.fromisoformat(tasks[0][9])).days

        ws_dashboard.cell(row=row, column=1, value=milestone_name)
        ws_dashboard.cell(row=row, column=2, value=day_offset)
        ws_dashboard.cell(row=row, column=3, value=date_str)
        row += 1

    # Create horizontal bar chart for milestone timeline (better visibility)
    chart_milestones = BarChart()
    chart_milestones.type = "bar"
    chart_milestones.title = "Project Milestones Timeline"
    chart_milestones.style = 26  # Colorful style
    chart_milestones.height = 10
    chart_milestones.width = 16
    chart_milestones.x_axis.title = "Day from Project Start"
    chart_milestones.y_axis.title = "Milestone Event"

    data = Reference(ws_dashboard, min_col=2, min_row=chart4_start_row + 1, max_row=row - 1)
    categories = Reference(ws_dashboard, min_col=1, min_row=chart4_start_row + 2, max_row=row - 1)
    chart_milestones.add_data(data, titles_from_data=True)
    chart_milestones.set_categories(categories)

    ws_dashboard.add_chart(chart_milestones, 'D42')

    # ========================================================================
    # Chart 5: Gantt-Style Timeline (Strategic Tasks Across Project)
    # ========================================================================
    chart5_start_row = chart4_start_row + 25
    ws_dashboard.cell(row=chart5_start_row, column=1, value="Gantt Timeline (Strategic Tasks by Phase)").font = Font(bold=True, size=11)
    ws_dashboard.cell(row=chart5_start_row + 1, column=1, value="Task").font = Font(bold=True)
    ws_dashboard.cell(row=chart5_start_row + 1, column=2, value="Phase").font = Font(bold=True)
    ws_dashboard.cell(row=chart5_start_row + 1, column=3, value="Start (Day)").font = Font(bold=True)
    ws_dashboard.cell(row=chart5_start_row + 1, column=4, value="Duration (Days)").font = Font(bold=True)

    # Select strategic tasks: longest task from each phase + overall longest tasks
    phase_longest = {}
    for task in tasks:
        phase = task[3]
        duration = task[8]
        if phase not in phase_longest or duration > phase_longest[phase][8]:
            phase_longest[phase] = task

    # Get strategic tasks (one per phase + top 5 longest overall)
    strategic_tasks = list(phase_longest.values())
    top_longest = sorted(tasks, key=lambda x: x[8], reverse=True)[:5]

    # Combine and remove duplicates
    strategic_set = {task[0]: task for task in strategic_tasks}  # Use task_id as key
    for task in top_longest:
        strategic_set[task[0]] = task

    # Sort by start date for chronological Gantt view
    selected_tasks = sorted(strategic_set.values(), key=lambda x: x[9])[:15]  # Limit to 15 for readability

    # Define phase colors (vibrant)
    phase_colors = {
        "Structure": "4472C4",      # Blue
        "Architecture": "ED7D31",   # Orange
        "MEP": "70AD47",            # Green
        "Finishes": "FFC000",       # Gold
        "Fit-out": "5B9BD5",        # Light Blue
        "Foundation": "A5A5A5",     # Gray
        "Commissioning": "C55A11",  # Brown
    }

    # Assign colors to phases present in data
    unique_phases = list(set(task[3] for task in selected_tasks))
    default_colors = ["4472C4", "ED7D31", "70AD47", "FFC000", "5B9BD5", "A5A5A5", "C55A11", "264478"]
    for i, phase in enumerate(unique_phases):
        if phase not in phase_colors:
            phase_colors[phase] = default_colors[i % len(default_colors)]

    row = chart5_start_row + 2
    for task in selected_tasks:
        task_name = task[2][:35]  # Truncate long names
        phase = task[3]
        start_date = datetime.fromisoformat(task[9])
        start_offset = (start_date - datetime.fromisoformat(tasks[0][9])).days
        duration = task[8]

        ws_dashboard.cell(row=row, column=1, value=task_name)
        ws_dashboard.cell(row=row, column=2, value=phase)
        ws_dashboard.cell(row=row, column=3, value=start_offset)
        ws_dashboard.cell(row=row, column=4, value=round(duration, 1))

        # Color-code phase cell
        phase_fill = PatternFill(start_color=phase_colors.get(phase, "DDDDDD"),
                                 end_color=phase_colors.get(phase, "DDDDDD"),
                                 fill_type="solid")
        ws_dashboard.cell(row=row, column=2).fill = phase_fill
        ws_dashboard.cell(row=row, column=2).font = Font(bold=True, color="FFFFFF")

        row += 1

    # Create stacked bar chart (Gantt-style)
    chart_gantt = BarChart()
    chart_gantt.type = "bar"
    chart_gantt.grouping = "stacked"
    chart_gantt.title = "Gantt Timeline - Strategic Tasks (Colored by Phase)"
    chart_gantt.style = 11
    chart_gantt.height = 12
    chart_gantt.width = 18
    chart_gantt.x_axis.title = "Days from Project Start"
    chart_gantt.y_axis.title = "Task"

    # Start offset (invisible base)
    start_data = Reference(ws_dashboard, min_col=3, min_row=chart5_start_row + 1, max_row=chart5_start_row + len(selected_tasks) + 1)
    # Duration (visible bars - color-coded)
    duration_data = Reference(ws_dashboard, min_col=4, min_row=chart5_start_row + 1, max_row=chart5_start_row + len(selected_tasks) + 1)
    categories = Reference(ws_dashboard, min_col=1, min_row=chart5_start_row + 2, max_row=chart5_start_row + len(selected_tasks) + 1)

    chart_gantt.add_data(start_data, titles_from_data=True)
    chart_gantt.add_data(duration_data, titles_from_data=True)
    chart_gantt.set_categories(categories)

    # Make start offset invisible (creates the "floating bar" Gantt effect)
    from openpyxl.chart.shapes import GraphicalProperties
    chart_gantt.series[0].graphicalProperties = GraphicalProperties()
    chart_gantt.series[0].graphicalProperties.noFill = True

    ws_dashboard.add_chart(chart_gantt, 'D58')

    # Format column widths
    ws_dashboard.column_dimensions['A'].width = 40
    ws_dashboard.column_dimensions['B'].width = 18
    ws_dashboard.column_dimensions['C'].width = 15
    ws_dashboard.column_dimensions['D'].width = 15

    # Save workbook
    wb.save(output_path)

    print(f"\n{'='*80}")
    print(f"✓ BIM 5D EXCEL SCHEDULE EXPORTED SUCCESSFULLY!")
    print(f"{'='*80}")
    print(f"  File: {output_path}")
    print(f"  Tasks: {len(tasks)}")
    print(f"\n  📊 Sheets:")
    print(f"    1. Construction Schedule - Full task list")
    print(f"    2. Project Summary - Phase/Discipline statistics")
    print(f"    3. BIM 5D Dashboard - Advanced analytics")
    print(f"\n  📈 Dashboard Charts (7 visualizations):")
    print(f"    • Phase Duration Analysis (Horizontal Bar)")
    print(f"    • Resource Workload by Phase (Column Chart)")
    print(f"    • S-Curve Progress (Line Chart)")
    print(f"    • Milestone Timeline (Scatter Plot)")
    print(f"    • Gantt Timeline - Top 15 Tasks (Stacked Bar)")
    print(f"    • Phase Distribution (Pie Chart)")
    print(f"    • Discipline Task Count (Bar Chart)")
    print(f"{'='*80}")
    print(f"  Open with: Excel | LibreOffice Calc | Google Sheets")
    print(f"{'='*80}\n")

    return output_path


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python excel_export.py <database_path> [output_xlsx_path] [project_name]")
        print("Example: python excel_export.py ~/Documents/bonsai/Terminal1_ARC_STR.db")
        sys.exit(1)

    db_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None
    project_name = sys.argv[3] if len(sys.argv) > 3 else "Terminal 1 Construction"

    export_schedule_to_excel(db_path, output_path, project_name)
