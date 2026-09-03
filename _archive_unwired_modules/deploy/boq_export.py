"""
Comprehensive BOQ Export - Malaysian Construction Industry Standard
Separate costing: Materials, Labor, Equipment with professional assumptions
Based on CIDB Malaysia 2024, BCISM Cost Book, and market research
"""

import sqlite3
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference

# ============================================================================
# MALAYSIAN CONSTRUCTION PRICING DATABASE (2024 Market Rates)
# Sources: CIDB N3C, BCISM Cost Book 2022-2024, Market Survey
# ============================================================================

# MATERIAL COSTS (RM per unit) - Updated for 2024 inflation (+3% CIDB forecast)
MATERIAL_COSTS = {
    # HVAC Ductwork (Galvanized Steel) - RM per meter
    'IfcDuct': {'rate': 165.00, 'unit': 'M', 'desc': 'Galvanized Steel Ductwork (avg 400mm)', 'spec': 'G550, 0.6mm thickness'},
    'IfcDuctSegment': {'rate': 165.00, 'unit': 'M', 'desc': 'Ductwork Segment', 'spec': 'G550, 0.6mm thickness'},
    'IfcDuctFitting': {'rate': 380.00, 'unit': 'EA', 'desc': 'Duct Fittings (elbows, tees)', 'spec': 'Galvanized steel'},

    # Plumbing (PVC/HDPE Pipes) - RM per meter
    'IfcPipe': {'rate': 48.50, 'unit': 'M', 'desc': 'PVC/HDPE Pipe (avg 100mm)', 'spec': 'Schedule 40, Class E'},
    'IfcPipeSegment': {'rate': 48.50, 'unit': 'M', 'desc': 'Pipe Segment', 'spec': 'Schedule 40'},
    'IfcPipeFitting': {'rate': 95.00, 'unit': 'EA', 'desc': 'Pipe Fittings', 'spec': 'PVC/brass mix'},

    # Electrical - RM per meter
    'IfcCableCarrier': {'rate': 78.00, 'unit': 'M', 'desc': 'Cable Tray System (300mm)', 'spec': 'Aluminum ladder type'},
    'IfcCableCarrierSegment': {'rate': 78.00, 'unit': 'M', 'desc': 'Cable Tray Segment', 'spec': 'Aluminum'},

    # Structural Steel - RM per tonne (converted to RM/M based on avg weight)
    # Mild steel: RM 3,800/tonne (2024 market rate)
    # Airport Terminal: Heavy sections, higher fabrication standards
    # I-Beam 400x200: ~66kg/m → RM 250/m material + RM 300/m fabrication
    # Column 400x400: ~146kg/m → RM 555/m material + RM 450/m fabrication
    'IfcBeam': {'rate': 680.00, 'unit': 'M', 'desc': 'Structural Steel I-Beam - Airport Spec', 'spec': 'Grade 50, shop fabrication, fire protection'},
    'IfcColumn': {'rate': 1250.00, 'unit': 'M', 'desc': 'Structural Steel Column - Airport Spec', 'spec': 'Grade 50, UC section, fire protection, heavy duty'},

    # Concrete & Slabs - RM per m² (AIRPORT TERMINAL SPEC - Higher grade)
    'IfcSlab': {'rate': 285.00, 'unit': 'M2', 'desc': 'RC Slab - Airport Grade (250mm)', 'spec': 'Grade 40 concrete, heavy rebar, formwork, curing'},

    # Walls - RM per m²
    'IfcWall': {'rate': 145.00, 'unit': 'M2', 'desc': 'Blockwork Wall (150mm)', 'spec': 'Cement block, plastered one side'},
    'IfcWallStandardCase': {'rate': 145.00, 'unit': 'M2', 'desc': 'Standard Wall', 'spec': 'Blockwork'},
    'IfcCurtainWall': {'rate': 750.00, 'unit': 'M2', 'desc': 'Aluminum Curtain Wall', 'spec': 'Double glazed, powder coated'},

    # Finishes - RM per m² (AIRPORT TERMINAL SPEC - Higher grade)
    'IfcCovering': {'rate': 185.00, 'unit': 'M2', 'desc': 'Floor/Ceiling Finishes - Airport Grade', 'spec': 'Granite tiles 600x600 / Metal suspended ceiling'},
    'IfcRoof': {'rate': 238.00, 'unit': 'M2', 'desc': 'Metal Deck Roofing with Insulation', 'spec': 'Standing seam, insulation, waterproofing'},

    # Fixtures - RM per piece (AIRPORT TERMINAL SPEC)
    'IfcLightFixture': {'rate': 485.00, 'unit': 'EA', 'desc': 'LED Light Fixture - Airport Grade', 'spec': 'Commercial 48W, dimming control, emergency backup'},
    'IfcOutlet': {'rate': 125.00, 'unit': 'EA', 'desc': '13A Power Outlet - Airport Spec', 'spec': 'Stainless steel plate, USB charging ports'},
    'IfcDoor': {'rate': 2850.00, 'unit': 'EA', 'desc': 'Door Set - Airport Grade', 'spec': 'Fire-rated, access control, automatic closer'},
    'IfcWindow': {'rate': 1580.00, 'unit': 'EA', 'desc': 'Aluminum Window - Airport Spec', 'spec': 'Powder coated, 12mm double glazed, acoustic'},
    'IfcBuildingElementProxy': {'rate': 850.00, 'unit': 'EA', 'desc': 'Misc. Building Elements', 'spec': 'Various fittings, furniture, signage'},
    'IfcFlowTerminal': {'rate': 3500.00, 'unit': 'EA', 'desc': 'HVAC Terminal (FCU/AHU/VAV)', 'spec': 'BMS integrated, VFD controls'},
}

# LABOR PRODUCTIVITY & RATES (Based on CIDB/MBAM Wage Survey 2024)
# Rates: Basic wage per day (8 hours) + 30% for EPF, SOCSO, benefits
LABOR_RATES = {
    # HVAC Installation
    'HVAC_TECH': {
        'rate_per_day': 185.00,  # RM per day (skilled)
        'productivity': {  # Output per crew per day
            'IfcDuct': 18.0,  # meters/day
            'IfcDuctSegment': 18.0,
            'IfcDuctFitting': 12.0,  # pieces/day
        },
        'crew_size': 2,  # Standard gang
        'trade': 'HVAC Technician (Skilled)'
    },

    # Plumbing Installation
    'PLUMBER': {
        'rate_per_day': 165.00,
        'productivity': {
            'IfcPipe': 25.0,  # meters/day
            'IfcPipeSegment': 25.0,
            'IfcPipeFitting': 15.0,  # pieces/day
        },
        'crew_size': 2,
        'trade': 'Pipefitter (Skilled)'
    },

    # Electrical Installation
    'ELECTRICIAN': {
        'rate_per_day': 175.00,
        'productivity': {
            'IfcCableCarrier': 30.0,  # meters/day
            'IfcCableCarrierSegment': 30.0,
            'IfcLightFixture': 20.0,  # pieces/day
            'IfcOutlet': 25.0,
        },
        'crew_size': 2,
        'trade': 'Electrician (Skilled)'
    },

    # Steel Erection
    'STEEL_ERECTOR': {
        'rate_per_day': 195.00,
        'productivity': {
            'IfcBeam': 8.0,  # meters/day (with crane)
            'IfcColumn': 6.0,
        },
        'crew_size': 4,  # Larger gang for safety
        'trade': 'Steel Erector (Skilled)'
    },

    # Concrete Works
    'CONCRETE_GANG': {
        'rate_per_day': 145.00,  # Average rate (mixed skilled/semi-skilled)
        'productivity': {
            'IfcSlab': 35.0,  # m²/day
        },
        'crew_size': 6,  # Formwork + concrete + finishing
        'trade': 'Concrete Gang (Mixed)'
    },

    # Masonry/Blockwork
    'MASON': {
        'rate_per_day': 155.00,
        'productivity': {
            'IfcWall': 12.0,  # m²/day
            'IfcWallStandardCase': 12.0,
        },
        'crew_size': 3,  # Mason + 2 laborers
        'trade': 'Mason (Skilled) + Laborers'
    },

    # General Labor (helpers)
    'LABORER': {
        'rate_per_day': 95.00,
        'crew_size': 1,
        'trade': 'General Laborer'
    },
}

# EQUIPMENT/MACHINERY HIRE RATES (RM per day)
# Based on CIDB N3C Machinery Hire Rates 2024
EQUIPMENT_RATES = {
    # Heavy lifting
    'MOBILE_CRANE_20T': {
        'rate_per_day': 1850.00,
        'description': 'Mobile Crane 20 Tonne',
        'operator_included': True,
        'fuel_included': False
    },
    'TOWER_CRANE': {
        'rate_per_day': 2200.00,
        'description': 'Tower Crane (monthly rental ÷ 26 days)',
        'operator_included': True,
        'fuel_included': True
    },

    # Concrete
    'CONCRETE_MIXER_10HP': {
        'rate_per_day': 85.00,
        'description': 'Concrete Mixer 10HP',
        'operator_included': False,
        'fuel_included': False
    },
    'CONCRETE_PUMP': {
        'rate_per_day': 950.00,
        'description': 'Concrete Pump Truck',
        'operator_included': True,
        'fuel_included': False
    },

    # Access equipment
    'SCISSOR_LIFT_8M': {
        'rate_per_day': 285.00,
        'description': 'Scissor Lift 8m working height',
        'operator_included': False,
        'fuel_included': False
    },
    'SCAFFOLDING': {
        'rate_per_month': 12.00,  # RM per m² per month
        'description': 'Steel Scaffolding System',
        'operator_included': False,
        'fuel_included': False
    },

    # Power tools
    'WELDING_MACHINE': {
        'rate_per_day': 65.00,
        'description': 'Welding Machine 300A',
        'operator_included': False,
        'fuel_included': False
    },
    'GENERATOR_5KVA': {
        'rate_per_day': 95.00,
        'description': 'Generator 5KVA',
        'operator_included': False,
        'fuel_included': False
    },
}

# EQUIPMENT ALLOCATION RULES (which equipment for which work)
EQUIPMENT_ALLOCATION = {
    'IfcBeam': {'equipment': 'MOBILE_CRANE_20T', 'duration_factor': 0.5},  # Crane for 50% of work duration
    'IfcColumn': {'equipment': 'MOBILE_CRANE_20T', 'duration_factor': 0.5},
    'IfcSlab': {'equipment': 'CONCRETE_PUMP', 'duration_factor': 0.3},
    'IfcDuct': {'equipment': 'SCISSOR_LIFT_8M', 'duration_factor': 0.4},
    'IfcCableCarrier': {'equipment': 'SCISSOR_LIFT_8M', 'duration_factor': 0.3},
}


class ComprehensiveBOQExporter:
    """Full BOQ with Materials, Labor, Equipment breakdown"""

    def __init__(self, output_path: str = None):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.output_path = output_path or f"BOQ_Comprehensive_{timestamp}.xlsx"

        self.colors = {
            'title': '1F4E78',
            'material': '70AD47',  # Green
            'labor': 'FFC000',     # Orange
            'equipment': '5B9BD5',  # Blue
            'total': 'C00000',      # Red
            'header': '4472C4',
        }

        self.sheet_refs = {}

    def calculate_labor_cost(self, ifc_class: str, quantity: float):
        """Calculate labor cost and crew-days"""
        # Find appropriate labor trade
        trade_key = None
        productivity = None

        for trade, data in LABOR_RATES.items():
            if 'productivity' in data and ifc_class in data['productivity']:
                trade_key = trade
                productivity = data['productivity'][ifc_class]
                break

        if not trade_key or not productivity:
            return 0, 0, 0, ''

        labor_data = LABOR_RATES[trade_key]
        crew_size = labor_data['crew_size']
        rate_per_day = labor_data['rate_per_day']

        # Calculate days needed
        days_needed = quantity / productivity

        # Total labor cost = days × crew_size × rate_per_day
        total_labor_cost = days_needed * crew_size * rate_per_day

        return total_labor_cost, days_needed, crew_size, labor_data['trade']

    def calculate_equipment_cost(self, ifc_class: str, labor_days: float):
        """Calculate equipment cost based on labor duration"""
        if ifc_class not in EQUIPMENT_ALLOCATION:
            return 0, ''

        alloc = EQUIPMENT_ALLOCATION[ifc_class]
        equipment_key = alloc['equipment']
        duration_factor = alloc['duration_factor']

        equipment_data = EQUIPMENT_RATES[equipment_key]
        equipment_days = labor_days * duration_factor
        equipment_cost = equipment_days * equipment_data['rate_per_day']

        return equipment_cost, equipment_data['description']

    def create_cover_sheet(self, project_name: str):
        """Enhanced cover sheet"""
        ws = self.wb.create_sheet("Cover Sheet", 0)

        ws.merge_cells('A1:F1')
        title = ws['A1']
        title.value = "COMPREHENSIVE BILL OF QUANTITIES"
        title.font = Font(size=22, bold=True, color='FFFFFF')
        title.fill = PatternFill(start_color=self.colors['title'], end_color=self.colors['title'], fill_type='solid')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35

        ws.merge_cells('A2:F2')
        ws['A2'] = f"Project: {project_name}"
        ws['A2'].font = Font(size=14, bold=True)
        ws['A2'].alignment = Alignment(horizontal='center')

        row = 4
        content = [
            ("", "", ""),
            ("COST BREAKDOWN METHODOLOGY", "", ""),
            ("", "", ""),
            ("1. MATERIAL COSTS", "", ""),
            ("Source:", "CIDB National Construction Cost Centre (N3C) 2024", ""),
            ("Reference:", "BCISM Cost Book 2022-2024 (inflated +3%)", ""),
            ("Includes:", "Delivery, wastage allowance (5-10% by material type)", ""),
            ("", "", ""),
            ("2. LABOR COSTS", "", ""),
            ("Source:", "MBAM-CIDB Labour Wage Survey 2024", ""),
            ("Basis:", "Basic wage + 30% (EPF 13%, SOCSO 2%, benefits 15%)", ""),
            ("Productivity:", "CIDB productivity standards by trade", ""),
            ("Crew Composition:", "Skilled workers + helpers as per trade standards", ""),
            ("", "", ""),
            ("3. EQUIPMENT/PLANT HIRE", "", ""),
            ("Source:", "CIDB N3C Machinery Hire Rates 2024", ""),
            ("Allocation:", "Based on work type and duration requirements", ""),
            ("Rates:", "Per day (8 hours), operator cost included where stated", ""),
            ("", "", ""),
            ("COST SUMMARY STRUCTURE", "", ""),
            ("Sheet 1:", "Cover Sheet (this page)", ""),
            ("Sheet 2:", "Executive Summary - Total costs with charts", ""),
            ("Sheet 3:", "Materials Cost Summary", "Detailed material breakdown"),
            ("Sheet 4:", "Labor Cost Summary", "Crew allocation & man-days"),
            ("Sheet 5:", "Equipment Cost Summary", "Plant hire requirements"),
            ("Sheet 6+:", "Detailed BOQ by Discipline", "Line-by-line analysis"),
            ("", "", ""),
            ("PRICING STANDARDS & REFERENCES", "", ""),
            ("BOQ Format:", "PWD Form 203A Malaysia", ""),
            ("Measurement:", "SMM2 (Standard Method of Measurement)", ""),
            ("Pricing Date:", "Q4 2024", ""),
            ("Currency:", "Malaysian Ringgit (RM)", ""),
            ("Validity:", "60 days from date of issue", ""),
            ("", "", ""),
            ("EXCLUSIONS", "", ""),
            ("• GST/SST (apply as per prevailing tax law)", "", ""),
            ("• Preliminary & General items (add 8-12%)", "", ""),
            ("• Profit & attendance (add 10-15%)", "", ""),
            ("• Escalation beyond 60 days", "", ""),
            ("• Site-specific conditions not shown in drawings", "", ""),
        ]

        for label, value, note in content:
            if not label:
                row += 1
                continue

            if "METHODOLOGY" in label or "SUMMARY" in label or "STANDARDS" in label or "EXCLUSIONS" in label:
                ws.merge_cells(f'A{row}:F{row}')
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(size=13, bold=True, color='FFFFFF')
                ws[f'A{row}'].fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
                ws.row_dimensions[row].height = 22
            elif "." in label and label[0].isdigit():  # Section numbers
                ws.merge_cells(f'A{row}:F{row}')
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(size=12, bold=True, color='FFFFFF')
                ws[f'A{row}'].fill = PatternFill(start_color=self.colors['material'], end_color=self.colors['material'], fill_type='solid')
            elif label.startswith('•'):
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(size=10)
            else:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'C{row}'] = note
                ws[f'A{row}'].font = Font(bold=True if label in ['Source:', 'Reference:', 'Basis:'] else False, size=10)
                ws[f'C{row}'].font = Font(size=9, italic=True)

            row += 1

        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 35

        return ws

    def create_detailed_boq_sheet(self, db_path: str, discipline: str):
        """Create detailed BOQ with Materials + Labor + Equipment breakdown"""
        sheet_name = f"BOQ - {discipline}"
        ws = self.wb.create_sheet(sheet_name)

        # Title
        ws.merge_cells('A1:K1')
        ws['A1'] = f"DETAILED BOQ - {discipline.upper()}"
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['title'], end_color=self.colors['title'], fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25

        # Headers
        row = 3
        headers = ['Item', 'Description', 'Qty', 'UOM',
                   'Material\nRate (RM)', 'Material\nCost (RM)',
                   'Labor\nCost (RM)', 'Equipment\nCost (RM)',
                   'Total (RM)']

        colors = ['FFFFFF', 'FFFFFF', 'FFFFFF', 'FFFFFF',
                  self.colors['material'], self.colors['material'],
                  self.colors['labor'], self.colors['equipment'],
                  self.colors['total']]

        for col, (header, color) in enumerate(zip(headers, colors), start=1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True, color='FFFFFF', size=9)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(bottom=Side(style='medium'))

        ws.row_dimensions[row].height = 30
        row += 1

        # Query data
        conn = sqlite3.connect(db_path)
        cursor = conn.execute("""
            SELECT ifc_class, total_quantity, uom
            FROM simple_qto
            WHERE discipline = ?
            ORDER BY total_quantity DESC
        """, (discipline,))

        data = cursor.fetchall()
        conn.close()

        if not data:
            return ws, None

        item_no = 1
        data_start = row

        for ifc_class, qty, uom in data:
            # Material
            mat = MATERIAL_COSTS.get(ifc_class, {'rate': 0, 'unit': uom, 'desc': ifc_class})
            mat_rate = mat['rate']
            mat_cost = qty * mat_rate

            # Labor
            labor_cost, labor_days, crew_size, trade = self.calculate_labor_cost(ifc_class, qty)

            # Equipment
            equip_cost, equip_desc = self.calculate_equipment_cost(ifc_class, labor_days)

            # Total
            total_cost = mat_cost + labor_cost + equip_cost

            # Write row
            ws.cell(row, 1, f"{item_no:03d}")
            ws.cell(row, 2, mat['desc'])
            ws.cell(row, 3, qty).number_format = '#,##0.00'
            ws.cell(row, 4, uom)
            ws.cell(row, 5, mat_rate).number_format = '#,##0.00'
            ws.cell(row, 6).value = f"=C{row}*E{row}"
            ws.cell(row, 6).number_format = '#,##0.00'
            ws.cell(row, 7, labor_cost).number_format = '#,##0.00'
            ws.cell(row, 8, equip_cost).number_format = '#,##0.00'
            ws.cell(row, 9).value = f"=F{row}+G{row}+H{row}"
            ws.cell(row, 9).number_format = '#,##0.00'

            # Yellow highlight on editable material rate
            ws.cell(row, 5).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

            item_no += 1
            row += 1

        data_end = row - 1

        # TOTALS
        row += 1
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = f"TOTAL - {discipline}"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'A{row}'].alignment = Alignment(horizontal='right')

        for col, color in [(6, self.colors['material']), (7, self.colors['labor']),
                           (8, self.colors['equipment']), (9, self.colors['total'])]:
            cell = ws.cell(row, col)
            cell.value = f"=SUM({get_column_letter(col)}{data_start}:{get_column_letter(col)}{data_end})"
            cell.number_format = '#,##0.00'
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        ws.row_dimensions[row].height = 22

        # Column widths
        widths = [8, 40, 12, 8, 15, 18, 16, 18, 18]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = 'A4'

        return ws, row

    def create_executive_summary(self, db_path: str, disciplines: list):
        """Executive summary with charts"""
        ws = self.wb.create_sheet("Executive Summary", 1)

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = "EXECUTIVE SUMMARY"
        ws['A1'].font = Font(size=18, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['title'], end_color=self.colors['title'], fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # Headers
        row = 3
        headers = ['Discipline', 'Material (RM)', 'Labor (RM)', 'Equipment (RM)', 'TOTAL (RM)']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True, color='FFFFFF', size=10)
            cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        row += 1

        summary_start = row

        # Link to detail sheets
        for disc in disciplines:
            if disc not in self.sheet_refs:
                continue

            sheet_name = f"BOQ - {disc}"
            total_row = self.sheet_refs[disc]

            ws.cell(row, 1, disc)
            for col in range(2, 6):
                cell = ws.cell(row, col)
                detail_col = 6 + (col - 2)  # 6=Material, 7=Labor, 8=Equipment, 9=Total
                cell.value = f"='{sheet_name}'!{get_column_letter(detail_col)}{total_row}"
                cell.number_format = '#,##0.00'

            row += 1

        # Add provisional sums if present
        if 'PROVISIONAL' in self.sheet_refs:
            prov_row = self.sheet_refs['PROVISIONAL']
            ws.cell(row, 1, "FINISHES & FITTINGS")
            ws.cell(row, 1).font = Font(italic=True)

            # Provisional doesn't split M/L/E in same way, so link total to last column
            ws.cell(row, 2).value = ""  # Material
            ws.cell(row, 3).value = ""  # Labor
            ws.cell(row, 4).value = ""  # Equipment
            ws.cell(row, 5).value = f"='Provisional Sums'!G{prov_row}"
            ws.cell(row, 5).number_format = '#,##0.00'

            row += 1

        summary_end = row - 1

        # Grand Total
        row += 1
        ws.merge_cells(f'A{row}:A{row}')
        ws[f'A{row}'] = "GRAND TOTAL"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].alignment = Alignment(horizontal='right')

        for col in range(2, 6):
            cell = ws.cell(row, col)
            cell.value = f"=SUM({get_column_letter(col)}{summary_start}:{get_column_letter(col)}{summary_end})"
            cell.number_format = '#,##0.00'
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color=self.colors['total'], end_color=self.colors['total'], fill_type='solid')

        ws.row_dimensions[row].height = 25

        # Column widths
        for col, width in enumerate([18, 18, 18, 18, 20], start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Add charts
        # Pie Chart - Cost Breakdown
        pie = PieChart()
        pie.title = "Cost Breakdown by Discipline"
        pie.style = 10
        pie.height = 10
        pie.width = 16

        labels = Reference(ws, min_col=1, min_row=summary_start, max_row=summary_end)
        data = Reference(ws, min_col=5, min_row=summary_start-1, max_row=summary_end)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        ws.add_chart(pie, 'H3')

        # Bar Chart - Cost Components
        bar = BarChart()
        bar.type = "col"
        bar.style = 10
        bar.title = "Cost Components by Discipline"
        bar.y_axis.title = "Cost (RM)"
        bar.x_axis.title = "Discipline"
        bar.height = 10
        bar.width = 16

        data = Reference(ws, min_col=2, max_col=4, min_row=summary_start-1, max_row=summary_end)
        cats = Reference(ws, min_col=1, min_row=summary_start, max_row=summary_end)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        ws.add_chart(bar, 'H23')

        ws.freeze_panes = 'A4'

        return ws

    def create_material_summary(self, db_path: str, disciplines: list):
        """Material cost summary"""
        ws = self.wb.create_sheet("Material Summary")

        ws.merge_cells('A1:F1')
        ws['A1'] = "MATERIAL COST SUMMARY"
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['material'], end_color=self.colors['material'], fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 25

        row = 3
        headers = ['Discipline', 'IFC Class', 'Quantity', 'UOM', 'Unit Rate (RM)', 'Total Material (RM)']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        row += 1

        # Query all material data
        conn = sqlite3.connect(db_path)
        cursor = conn.execute("""
            SELECT discipline, ifc_class, total_quantity, uom
            FROM simple_qto
            ORDER BY discipline, total_quantity DESC
        """)

        for disc, ifc_class, qty, uom in cursor.fetchall():
            mat = MATERIAL_COSTS.get(ifc_class, {'rate': 0})
            mat_rate = mat['rate']
            mat_cost = qty * mat_rate

            ws.cell(row, 1, disc)
            ws.cell(row, 2, ifc_class)
            ws.cell(row, 3, qty).number_format = '#,##0.00'
            ws.cell(row, 4, uom)
            ws.cell(row, 5, mat_rate).number_format = '#,##0.00'
            ws.cell(row, 6, mat_cost).number_format = '#,##0.00'

            row += 1

        conn.close()

        # Column widths
        for col, width in enumerate([15, 28, 12, 8, 16, 20], start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = 'A4'
        return ws

    def create_labor_summary(self, db_path: str, disciplines: list):
        """Labor cost and crew allocation summary"""
        ws = self.wb.create_sheet("Labor Summary")

        ws.merge_cells('A1:H1')
        ws['A1'] = "LABOR COST & CREW ALLOCATION SUMMARY"
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['labor'], end_color=self.colors['labor'], fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 25

        row = 3
        headers = ['Discipline', 'IFC Class', 'Quantity', 'UOM', 'Trade', 'Crew Size', 'Man-Days', 'Labor Cost (RM)']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        row += 1

        # Query all data
        conn = sqlite3.connect(db_path)
        cursor = conn.execute("""
            SELECT discipline, ifc_class, total_quantity, uom
            FROM simple_qto
            ORDER BY discipline, total_quantity DESC
        """)

        for disc, ifc_class, qty, uom in cursor.fetchall():
            labor_cost, labor_days, crew_size, trade = self.calculate_labor_cost(ifc_class, qty)

            if labor_cost == 0:
                continue

            man_days = labor_days * crew_size

            ws.cell(row, 1, disc)
            ws.cell(row, 2, ifc_class)
            ws.cell(row, 3, qty).number_format = '#,##0.00'
            ws.cell(row, 4, uom)
            ws.cell(row, 5, trade)
            ws.cell(row, 6, crew_size)
            ws.cell(row, 7, man_days).number_format = '#,##0.0'
            ws.cell(row, 8, labor_cost).number_format = '#,##0.00'

            row += 1

        conn.close()

        # Column widths
        for col, width in enumerate([15, 28, 12, 8, 30, 10, 12, 18], start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = 'A4'
        return ws

    def create_equipment_summary(self, db_path: str, disciplines: list):
        """Equipment/plant hire summary"""
        ws = self.wb.create_sheet("Equipment Summary")

        ws.merge_cells('A1:G1')
        ws['A1'] = "EQUIPMENT / PLANT HIRE SUMMARY"
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['equipment'], end_color=self.colors['equipment'], fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 25

        row = 3
        headers = ['Discipline', 'IFC Class', 'Equipment', 'Duration (days)', 'Rate/Day (RM)', 'Total (RM)']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        row += 1

        # Query all data
        conn = sqlite3.connect(db_path)
        cursor = conn.execute("""
            SELECT discipline, ifc_class, total_quantity
            FROM simple_qto
            ORDER BY discipline, total_quantity DESC
        """)

        for disc, ifc_class, qty in cursor.fetchall():
            # Calculate labor days first
            labor_cost, labor_days, crew_size, trade = self.calculate_labor_cost(ifc_class, qty)

            # Calculate equipment
            equip_cost, equip_desc = self.calculate_equipment_cost(ifc_class, labor_days)

            if equip_cost == 0:
                continue

            # Get allocation details
            alloc = EQUIPMENT_ALLOCATION.get(ifc_class, {})
            equip_key = alloc.get('equipment', '')
            duration_factor = alloc.get('duration_factor', 0)
            equip_days = labor_days * duration_factor

            rate_per_day = EQUIPMENT_RATES.get(equip_key, {}).get('rate_per_day', 0)

            ws.cell(row, 1, disc)
            ws.cell(row, 2, ifc_class)
            ws.cell(row, 3, equip_desc)
            ws.cell(row, 4, equip_days).number_format = '#,##0.0'
            ws.cell(row, 5, rate_per_day).number_format = '#,##0.00'
            ws.cell(row, 6, equip_cost).number_format = '#,##0.00'

            row += 1

        conn.close()

        # Column widths
        for col, width in enumerate([15, 28, 35, 15, 16, 18], start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = 'A4'
        return ws

    def create_provisional_sums(self, db_path: str):
        """Additional finishes and fittings not in BIM model"""
        ws = self.wb.create_sheet("Provisional Sums")

        ws.merge_cells('A1:F1')
        ws['A1'] = "PROVISIONAL SUMS - FINISHES & FITTINGS"
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='8E44AD', end_color='8E44AD', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 25

        ws['A2'] = "Items not explicitly modeled in BIM - calculated from built-up areas"
        ws['A2'].font = Font(size=9, italic=True)
        ws.merge_cells('A2:F2')

        # Get total wall and floor areas
        conn = sqlite3.connect(db_path)

        cursor = conn.execute("""
            SELECT SUM(total_quantity) FROM simple_qto
            WHERE ifc_class IN ('IfcWall', 'IfcWallStandardCase')
        """)
        wall_area = cursor.fetchone()[0] or 0

        cursor = conn.execute("""
            SELECT SUM(total_quantity) FROM simple_qto
            WHERE ifc_class = 'IfcSlab'
        """)
        floor_area = cursor.fetchone()[0] or 0

        conn.close()

        # Provisional items with rates
        prov_items = [
            {
                'item': 'P1',
                'desc': 'Painting - Walls and Ceilings',
                'qty': wall_area * 2,  # Both sides
                'uom': 'M2',
                'mat_rate': 8.50,
                'labor_rate': 12.00,
                'spec': 'Dulux/Nippon 2 coats emulsion, primer, preparation'
            },
            {
                'item': 'P2',
                'desc': 'Check-in Counters & Service Desks',
                'qty': 15,  # Typical terminal
                'uom': 'EA',
                'mat_rate': 12500.00,
                'labor_rate': 2500.00,
                'spec': 'Solid surface, modular system, cable mgmt, branding'
            },
            {
                'item': 'P3',
                'desc': 'Passenger Seating - Waiting Areas',
                'qty': 250,  # Seats
                'uom': 'EA',
                'mat_rate': 580.00,
                'labor_rate': 45.00,
                'spec': 'Airport beam seating, 3-4 seater units, steel frame'
            },
            {
                'item': 'P4',
                'desc': 'Wayfinding Signage System',
                'qty': 80,  # Signs
                'uom': 'EA',
                'mat_rate': 650.00,
                'labor_rate': 120.00,
                'spec': 'Illuminated, bilingual, airport pictograms'
            },
            {
                'item': 'P5',
                'desc': 'Flight Information Display Systems (FIDS)',
                'qty': 25,  # Screens
                'uom': 'EA',
                'mat_rate': 8500.00,
                'labor_rate': 1200.00,
                'spec': '55" LED, networked, real-time updates'
            },
            {
                'item': 'P6',
                'desc': 'Baggage Trolley Storage Racks',
                'qty': 12,
                'uom': 'EA',
                'mat_rate': 1850.00,
                'labor_rate': 350.00,
                'spec': 'Stainless steel, capacity 30 trolleys each'
            },
            {
                'item': 'P7',
                'desc': 'Retail Kiosk Fit-outs (Shell)',
                'qty': 8,
                'uom': 'EA',
                'mat_rate': 22000.00,
                'labor_rate': 5500.00,
                'spec': 'Structural frame, services rough-in, ready for tenant'
            },
            {
                'item': 'P8',
                'desc': 'Rubber Safety Flooring - High Traffic',
                'qty': floor_area * 0.15,  # 15% of floor area
                'uom': 'M2',
                'mat_rate': 95.00,
                'labor_rate': 28.00,
                'spec': 'Anti-slip, heavy duty, transition zones'
            },
            {
                'item': 'P9',
                'desc': 'Acoustic Ceiling Panels - Special Areas',
                'qty': floor_area * 0.20,  # 20% special treatment
                'uom': 'M2',
                'mat_rate': 135.00,
                'labor_rate': 42.00,
                'spec': 'High absorption, Class A, fire-rated'
            },
            {
                'item': 'P10',
                'desc': 'Bollards & Barriers - Security',
                'qty': 45,
                'uom': 'EA',
                'mat_rate': 1250.00,
                'labor_rate': 280.00,
                'spec': 'Fixed/removable, stainless steel, crash-rated'
            },
        ]

        # Headers
        row = 4
        headers = ['Item', 'Description', 'Qty', 'UOM', 'Material\nRate (RM)', 'Labor\nRate (RM)', 'Total (RM)']
        colors = ['FFFFFF', 'FFFFFF', 'FFFFFF', 'FFFFFF', self.colors['material'], self.colors['labor'], self.colors['total']]

        for col, (header, color) in enumerate(zip(headers, colors), start=1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True, color='FFFFFF', size=9)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[row].height = 30
        row += 1

        data_start = row

        # Data rows
        for item in prov_items:
            ws.cell(row, 1, item['item'])
            ws.cell(row, 2, f"{item['desc']}\n({item['spec']})").alignment = Alignment(wrap_text=True)
            ws.cell(row, 3, item['qty']).number_format = '#,##0.00'
            ws.cell(row, 4, item['uom'])
            ws.cell(row, 5, item['mat_rate']).number_format = '#,##0.00'
            ws.cell(row, 6, item['labor_rate']).number_format = '#,##0.00'

            # Formula for total
            ws.cell(row, 7).value = f"=C{row}*(E{row}+F{row})"
            ws.cell(row, 7).number_format = '#,##0.00'

            # Yellow highlight on editable rates
            ws.cell(row, 5).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            ws.cell(row, 6).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

            ws.row_dimensions[row].height = 35
            row += 1

        data_end = row - 1

        # TOTAL
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "TOTAL PROVISIONAL SUMS"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'A{row}'].alignment = Alignment(horizontal='right')

        cell = ws.cell(row, 7)
        cell.value = f"=SUM(G{data_start}:G{data_end})"
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color=self.colors['total'], end_color=self.colors['total'], fill_type='solid')

        ws.row_dimensions[row].height = 25

        # Save total row for summary
        self.sheet_refs['PROVISIONAL'] = row

        # Column widths
        for col, width in enumerate([8, 48, 12, 8, 15, 15, 18], start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = 'A5'
        return ws

    def generate_comprehensive_boq(self, db_path: str, project_name: str):
        """Generate full BOQ"""
        print(f"\n{'='*80}")
        print("GENERATING COMPREHENSIVE BOQ")
        print(f"{'='*80}\n")

        # Get disciplines
        conn = sqlite3.connect(db_path)
        cursor = conn.execute("SELECT DISTINCT discipline FROM simple_qto ORDER BY discipline")
        disciplines = [r[0] for r in cursor.fetchall()]
        conn.close()

        print(f"Creating cover sheet...")
        self.create_cover_sheet(project_name)

        # Detail sheets
        for disc in disciplines:
            print(f"Creating BOQ - {disc}...")
            ws, total_row = self.create_detailed_boq_sheet(db_path, disc)
            if total_row:
                self.sheet_refs[disc] = total_row

        # Add provisional sums sheet for additional works
        print(f"\nCreating provisional sums for finishes...")
        self.create_provisional_sums(db_path)

        # Summary sheets
        print(f"\nCreating summary sheets...")
        self.create_executive_summary(db_path, disciplines)
        self.create_material_summary(db_path, disciplines)
        self.create_labor_summary(db_path, disciplines)
        self.create_equipment_summary(db_path, disciplines)

        print(f"\nSaving: {self.output_path}")
        self.wb.save(self.output_path)

        print(f"\n{'='*80}")
        print("✓ COMPREHENSIVE BOQ COMPLETE")
        print(f"{'='*80}\n")

        return self.output_path


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python comprehensive_boq_export.py <database_path> [project_name]")
        sys.exit(1)

    db_path = sys.argv[1]
    project_name = sys.argv[2] if len(sys.argv) > 2 else "Terminal 1 Expansion Project"

    exporter = ComprehensiveBOQExporter()
    exporter.generate_comprehensive_boq(db_path, project_name)
